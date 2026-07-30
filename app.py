from datetime import datetime, timezone, timedelta
from flask import Flask, request, jsonify
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity, decode_token
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import sys
import os
import pandas as pd
from flask import send_file
import openpyxl
from io import BytesIO
import tempfile
import math
import barcode
from barcode.writer import ImageWriter
import io
from flask import send_file
from markupsafe import escape

from database import get_connection, init_db, calcular_vlc_total, USE_POSTGRES

from functools import wraps

def registrar_log(usuario_id, usuario_nome, acao, tabela, descricao, conn=None):
    """Registra uma ação no log de auditoria."""
    try:
        # Se a função recebeu uma conexão já aberta (carona), usa ela:
        if conn:
            conn.execute(
                '''INSERT INTO log_auditoria (usuario_id, usuario_nome, acao, tabela, descricao)
                   VALUES (?, ?, ?, ?, ?)''',
                (usuario_id, usuario_nome, acao, tabela, descricao)
            )
        # Se não recebeu (ação isolada), abre uma nova conexão normalmente:
        else:
            with get_connection() as nova_conn:
                nova_conn.execute(
                    '''INSERT INTO log_auditoria (usuario_id, usuario_nome, acao, tabela, descricao)
                       VALUES (?, ?, ?, ?, ?)''',
                    (usuario_id, usuario_nome, acao, tabela, descricao)
                )
                nova_conn.commit()
    except Exception as e:
        print(f'Erro ao registrar log: {e}')

def perfil_required(*perfis_permitidos):
    """Decorator que verifica se o usuário tem um dos perfis permitidos."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            current_user = get_jwt_identity()
            with get_connection() as conn:
                user = conn.execute(
                    'SELECT id, perfil FROM usuarios WHERE username = ?',
                    (current_user,)
                ).fetchone()
                if not user or user['perfil'] not in perfis_permitidos:
                    return response(False, message='Acesso negado. Permissão insuficiente.', status_code=403)
            return f(*args, **kwargs)
        return decorated
    return decorator

app = Flask(__name__)
app.config['JWT_SECRET_KEY'] = 'super-secret-key-almoxarifado-2026'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=8)

jwt = JWTManager(app)
CORS(app, resources={r"/*": {"origins": "*"}})


def response(success=True, data=None, message=None, status_code=200, pagination=None):
    body = {'success': success}
    if data is not None:
        body['data'] = data
    if message is not None:
        body['message'] = message
    if pagination is not None:
        body['pagination'] = pagination
    return jsonify(body), status_code


def rows_to_dict(rows):
    return [dict(row) for row in rows]


def now_iso():
    # Trava o fuso horário no padrão de Brasília (UTC-3)
    fuso_br = timezone(timedelta(hours=-3))
    return datetime.now(fuso_br).isoformat()

def calcular_custo_medio(conn, produto_id, quantidade_entrada, valor_unitario):
    """Calcula o novo custo médio ponderado após uma entrada."""
    saldo_row = conn.execute(
        'SELECT COALESCE(SUM(quantidade), 0) as saldo FROM estoque WHERE produto_id = ?',
        (produto_id,)
    ).fetchone()
    saldo_atual = saldo_row['saldo'] if saldo_row else 0

    custo_row = conn.execute(
        'SELECT COALESCE(custo_medio, 0) as custo_medio FROM produtos WHERE id = ?',
        (produto_id,)
    ).fetchone()
    custo_medio_antigo = custo_row['custo_medio'] if custo_row else 0

    if saldo_atual == 0:
        return valor_unitario

    novo_custo = (saldo_atual * custo_medio_antigo + quantidade_entrada * valor_unitario) / (saldo_atual + quantidade_entrada)
    return novo_custo

# ============================================================
# AUTENTICAÇÃO
# ============================================================

@app.route('/registro', methods=['POST'])
@jwt_required()
@perfil_required('admin')
def registro():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    perfil = data.get('perfil', 'operador')

    if not username or not password:
        return response(False, message='Username e password são obrigatórios.', status_code=400)

    if perfil not in ('admin', 'operador', 'visualizador'):
        return response(False, message='Perfil inválido. Use admin, operador ou visualizador.', status_code=400)

    try:
        with get_connection() as conn:
            cur = conn.execute('SELECT id FROM usuarios WHERE username = ?', (username,))
            if cur.fetchone():
                return response(False, message='Username já existe.', status_code=409)

            conn.execute(
                'INSERT INTO usuarios (username, password, perfil, created_at) VALUES (?, ?, ?, ?)',
                (username, generate_password_hash(password), perfil, now_iso())
            )
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            registrar_log(user['id'], current_user, 'criar', 'usuarios', f'Criou usuário "{username}" com perfil {perfil}')

        return response(True, message='Usuário registrado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)
    
@app.route('/usuarios', methods=['GET'])
@jwt_required()
@perfil_required('admin')
def listar_usuarios():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                'SELECT id, username, perfil, created_at FROM usuarios ORDER BY username'
            ).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)
    
@app.route('/usuarios/<int:id>', methods=['GET'])
@jwt_required()
@perfil_required('admin')
def get_usuario(id):
    try:
        with get_connection() as conn:
            row = conn.execute(
                'SELECT id, username, email, perfil, ativo FROM usuarios WHERE id = ?',
                (id,)
            ).fetchone()

            if not row:
                return response(False, message='Usuário não encontrado.', status_code=404)

            user = dict(row)
            user['ativo'] = bool(user['ativo'])

        return response(True, data=user)
    except Exception as e:
        return response(False, message=str(e), status_code=500)    

@app.route('/usuarios/<int:id>', methods=['PUT'])
@jwt_required()
@perfil_required('admin')
def atualizar_usuario(id):
    data = request.get_json() or {}
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    perfil = data.get('perfil')
    ativo = data.get('ativo')

    if perfil and perfil not in ('admin', 'operador', 'visualizador'):
        return response(False, message='Perfil inválido.', status_code=400)

    try:
        with get_connection() as conn:
            # Monta os campos dinamicamente
            updates = []
            params = []

            if username:
                updates.append('username = ?')
                params.append(username)
            if email is not None:
                updates.append('email = ?')
                params.append(email)
            if password:
                hashed = generate_password_hash(password)
                updates.append('password = ?')
                params.append(hashed)
            if perfil:
                updates.append('perfil = ?')
                params.append(perfil)
            if ativo is not None:
                updates.append('ativo = ?')
                params.append(1 if ativo else 0)

            if not updates:
                return response(False, message='Nenhum campo para atualizar.', status_code=400)

            params.append(id)
            conn.execute(f'UPDATE usuarios SET {", ".join(updates)} WHERE id = ?', params)
            conn.commit()

            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            campos = ', '.join(u.split(' = ')[0] for u in updates)
            registrar_log(user['id'], current_user, 'editar', 'usuarios', f'Atualizou ({campos}) do usuário ID {id}')

        return response(True, message='Usuário atualizado com sucesso!')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/usuarios/<int:id>', methods=['DELETE'])
@jwt_required()
@perfil_required('admin')
def excluir_usuario(id):
    try:
        with get_connection() as conn:
            row = conn.execute('SELECT username FROM usuarios WHERE id = ?', (id,)).fetchone()
            if not row:
                return response(False, message='Usuário não encontrado.', status_code=404)

            conn.execute('DELETE FROM usuarios WHERE id = ?', (id,))
            conn.commit()

            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            registrar_log(user['id'], current_user, 'excluir', 'usuarios', f'Excluiu usuário "{row["username"]}" (ID {id})')

        return response(True, message='Usuário excluído.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/log-auditoria', methods=['GET'])
@jwt_required()
@perfil_required('admin', 'operador')
def listar_log():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                '''SELECT * FROM log_auditoria ORDER BY data DESC LIMIT 500'''
            ).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)              

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return response(False, message='Username e password são obrigatórios.', status_code=400)

    try:
        with get_connection() as conn:
            row = conn.execute(
                'SELECT * FROM usuarios WHERE username = ?', (username,)
            ).fetchone()

        if not row or not check_password_hash(row['password'], password):
            return response(False, message='Credenciais inválidas.', status_code=401)

        access_token = create_access_token(identity=username)
        return response(True, data={
            'access_token': access_token,
            'perfil': row['perfil'] if 'perfil' in row.keys() else 'admin',
            'usuario_id': row['id'],
            'username': row['username']
        })
    except Exception as e:
        return response(False, message=str(e), status_code=500)


# ============================================================
# PRODUTOS
# ============================================================

@app.route('/produtos', methods=['GET'])
@jwt_required()
def listar_produtos():
    try:
        with get_connection() as conn:
            rows = conn.execute('''
                SELECT p.*, a.nome as almoxarifado_nome,
                    COALESCE((SELECT SUM(quantidade) FROM estoque WHERE produto_id = p.id), 0) as saldo_total
                FROM produtos p
                LEFT JOIN almoxarifados a ON p.almoxarifado_id = a.id
                ORDER BY
                    CASE
                        WHEN COALESCE((SELECT SUM(quantidade) FROM estoque WHERE produto_id = p.id), 0) <= 0 THEN 0
                        WHEN COALESCE((SELECT SUM(quantidade) FROM estoque WHERE produto_id = p.id), 0) < COALESCE(p.estoque_minimo, 0) THEN 1
                        ELSE 2
                    END,
                    p.nome
            ''').fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/produtos', methods=['POST'])
@jwt_required()
def criar_produto():
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()

    if not nome:
        return response(False, message='Nome é obrigatório.', status_code=400)
    
    campos = [
        'nome', 'codigo_interno', 'codigo_fabricante', 'codigo_barras',
        'descricao', 'categoria', 'tipo', 'unidade', 'preco', 'valor_unitario',
        'estoque_minimo', 'sobressalente', 'rastreabilidade', 'controla_depreciacao',
        'requer_equipamento', 'equipamentos_compativeis'
    ]

    valores = {campo: data.get(campo) for campo in campos}
    valores['nome'] = nome
    valores['controla_depreciacao'] = bool(valores.get('controla_depreciacao', 0))
    valores['requer_equipamento'] = bool(valores.get('requer_equipamento', 0))
    valores['created_at'] = now_iso()
    valores['updated_at'] = now_iso()

    colunas = ', '.join(valores.keys())
    placeholders = ', '.join(['?' for _ in valores])

    try:
        with get_connection() as conn:
            conn.execute(
                f'INSERT INTO produtos ({colunas}) VALUES ({placeholders})',
                tuple(valores.values())
            )
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'criar', 'produtos',
                              f'Criou produto "{nome}"')

        return response(True, message='Produto criado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/produtos/<int:id>', methods=['PUT'])
@jwt_required()
def atualizar_produto(id):
    data = request.get_json() or {}

    campos = [
        'nome', 'codigo_interno', 'codigo_fabricante', 'codigo_barras',
        'descricao', 'categoria', 'tipo', 'unidade', 'preco', 'valor_unitario',
        'estoque_minimo', 'sobressalente', 'rastreabilidade', 'controla_depreciacao',
        'requer_equipamento', 'equipamentos_compativeis'
    ]

    # ✅ Só atualiza os campos que vieram na requisição
    valores = {}
    for campo in campos:
        if campo in data:
            valor = data[campo]
            if campo == 'nome':
                valor = valor.strip() if valor else ''
            valores[campo] = valor

        # Converte campos booleanos para compatibilidade com PostgreSQL
    if 'controla_depreciacao' in valores:
        valores['controla_depreciacao'] = bool(valores['controla_depreciacao'])
    if 'requer_equipamento' in valores:
        valores['requer_equipamento'] = bool(valores['requer_equipamento'])

    # ✅ Valida nome apenas se foi enviado
    if 'nome' in valores and not valores['nome']:
        return response(False, message='Nome não pode ser vazio.', status_code=400)

    if not valores:
        return response(False, message='Nenhum campo para atualizar.', status_code=400)

    valores['updated_at'] = now_iso()

    set_clause = ', '.join([f'{campo} = ?' for campo in valores.keys()])
    params = tuple(valores.values()) + (id,)

    try:
        with get_connection() as conn:
            cur = conn.execute(
                f'UPDATE produtos SET {set_clause} WHERE id = ?', params
            )

            if cur.rowcount == 0:
                return response(False, message='Produto não encontrado.', status_code=404)

            # Log de auditoria
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()

            if user:
                prod = conn.execute(
                    'SELECT nome FROM produtos WHERE id = ?', (id,)
                ).fetchone()
                nome_log = prod['nome'] if prod else f'ID {id}'
                registrar_log(
                    user['id'],
                    current_user,
                    'editar',
                    'produtos',
                    f'Atualizou o produto ID {id} ({nome_log})',
                    conn=conn
                )

            conn.commit()

        return response(True, message='Produto atualizado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/produtos/<int:id>', methods=['DELETE'])
@jwt_required()
def excluir_produto(id):
    try:
        with get_connection() as conn:
            # Primeiro busca o nome do produto para o log
            row = conn.execute('SELECT nome FROM produtos WHERE id = ?', (id,)).fetchone()
            if not row:
                return response(False, message='Produto não encontrado.', status_code=404)

            conn.execute('DELETE FROM produtos WHERE id = ?', (id,))
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'excluir', 'produtos',
                              f'Excluiu produto "{row["nome"]}" (ID {id})')

        return response(True, message='Produto excluído com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/produtos/importar', methods=['POST'])
@jwt_required()
def importar_produtos():
    data = request.get_json() or {}
    produtos = data.get('produtos', [])

    if not isinstance(produtos, list):
        return response(False, message='O campo produtos deve ser uma lista.', status_code=400)

    try:
        with get_connection() as conn:
            importados = 0
            atualizados = 0
            
            for p in produtos:
                nome = p.get('nome', '').strip()
                if not nome:
                    continue
                    
                # Pega o código interno (ou fabricante, dependendo de como veio da planilha)
                codigo_interno = p.get('codigo_interno', '').strip()
                if not codigo_interno:
                    codigo_interno = p.get('codigo_fabricante', '').strip()

                # 1. VERIFICA SE O PRODUTO JÁ EXISTE
                produto_existente = None
                if codigo_interno:
                    # Busca pelo código
                    produto_existente = conn.execute(
                        'SELECT id FROM produtos WHERE codigo_interno = ? OR codigo_fabricante = ?', 
                        (codigo_interno, codigo_interno)
                    ).fetchone()
                else:
                    # Se o produto não tem código na planilha, busca pelo nome exato para não duplicar
                    produto_existente = conn.execute(
                        'SELECT id FROM produtos WHERE nome = ?', 
                        (nome,)
                    ).fetchone()

                # 2. PREPARA OS VALORES
                codigo_fabricante = p.get('codigo_fabricante', '')
                codigo_barras = p.get('codigo_barras', '')
                descricao = p.get('descricao', '')
                categoria = p.get('categoria', '')
                tipo = p.get('tipo', '')
                unidade = p.get('unidade', '')
                valor_real = p.get('valor_unitario') or p.get('preco') or 0
                preco = valor_real
                valor_unitario = valor_real
                estoque_minimo = p.get('estoque_minimo', 0)
                rastreabilidade = p.get('rastreabilidade', '')
                almoxarifado_id = p.get('almoxarifado_id', None)
                
                agora = now_iso()

                # 3. DECIDE SE ATUALIZA OU CRIA NOVO
                if produto_existente:
                    # UPDATE: Atualiza o produto que já existe (inclusive o preço corrigido!)
                    produto_id = produto_existente['id']
                    conn.execute('''
                        UPDATE produtos 
                        SET nome = ?, codigo_interno = ?, codigo_fabricante = ?, codigo_barras = ?,
                            descricao = ?, categoria = ?, tipo = ?, unidade = ?, preco = ?, 
                            valor_unitario = ?, estoque_minimo = ?, rastreabilidade = ?, 
                            almoxarifado_id = ?, updated_at = ?
                        WHERE id = ?
                    ''', (
                        nome, codigo_interno, codigo_fabricante, codigo_barras,
                        descricao, categoria, tipo, unidade, preco,
                        valor_unitario, estoque_minimo, rastreabilidade,
                        almoxarifado_id, agora, produto_id
                    ))
                    atualizados += 1
                else:
                    # INSERT: Cria um produto totalmente novo
                    conn.execute('''
                        INSERT INTO produtos (
                            nome, codigo_interno, codigo_fabricante, codigo_barras,
                            descricao, categoria, tipo, unidade, preco, valor_unitario,
                            estoque_minimo, rastreabilidade, almoxarifado_id, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        nome, codigo_interno, codigo_fabricante, codigo_barras,
                        descricao, categoria, tipo, unidade, preco, valor_unitario,
                        estoque_minimo, rastreabilidade, almoxarifado_id, agora, agora
                    ))
                    importados += 1
                    
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'importar', 'produtos',
                              f'Planilha: {importados} criados, {atualizados} atualizados')

        # Retorna a mensagem detalhada para o usuário
        mensagem_final = f'Sucesso! {importados} novos criados e {atualizados} atualizados.'
        return response(True, message=mensagem_final, data={'importados': importados, 'atualizados': atualizados})
        
    except Exception as e:
        return response(False, message=str(e), status_code=500)


# ============================================================
# ALMOXARIFADOS
# ============================================================

@app.route('/almoxarifados', methods=['GET'])
@jwt_required()
def listar_almoxarifados():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                'SELECT * FROM almoxarifados ORDER BY nome'
            ).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/almoxarifados', methods=['POST'])
@jwt_required()
def criar_almoxarifado():
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()
    codigo = data.get('codigo', '').strip()

    if not nome:
        return response(False, message='Nome é obrigatório.', status_code=400)
    if not codigo:
        return response(False, message='Código é obrigatório.', status_code=400)

    valores = {
        'nome': nome,
        'codigo': codigo,
        'responsavel': data.get('responsavel'),
        'localizacao': data.get('localizacao'),
        'descricao': data.get('descricao'),
        'status': data.get('status') or 'ativo',
        'created_at': now_iso(),
        'updated_at': now_iso()
    }
    colunas = ', '.join(valores.keys())
    placeholders = ', '.join(['?' for _ in valores])

    try:
        with get_connection() as conn:
            conn.execute(
                f'INSERT INTO almoxarifados ({colunas}) VALUES ({placeholders})',
                tuple(valores.values())
            )
            conn.commit()

            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'criar', 'almoxarifados',
                              f'Criou almoxarifado "{nome}"')

        return response(True, message='Almoxarifado criado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/almoxarifados/<int:id>', methods=['PUT'])
@jwt_required()
def atualizar_almoxarifado(id):
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()

    if not nome:
        return response(False, message='Nome é obrigatório.', status_code=400)

    valores = {
        'nome': nome,
        'codigo': data.get('codigo'),
        'responsavel': data.get('responsavel'),
        'localizacao': data.get('localizacao'),
        'descricao': data.get('descricao'),
        'status': data.get('status'),
        'updated_at': now_iso()
    }

    set_clause = ', '.join([f'{campo} = ?' for campo in valores.keys()])
    params = tuple(valores.values()) + (id,)

    try:
        with get_connection() as conn:
            cur = conn.execute(
                f'UPDATE almoxarifados SET {set_clause} WHERE id = ?', params
            )
            conn.commit()
            if cur.rowcount == 0:
                return response(False, message='Almoxarifado não encontrado.', status_code=404)
        return response(True, message='Almoxarifado atualizado com sucesso.')
    except sqlite3.IntegrityError:
        return response(False, message='Código do almoxarifado já existe.', status_code=409)
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/almoxarifados/<int:id>', methods=['DELETE'])
@jwt_required()
def excluir_almoxarifado(id):
    try:
        with get_connection() as conn:
            # Primeiro busca o nome para o log
            row = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (id,)).fetchone()
            if not row:
                return response(False, message='Almoxarifado não encontrado.', status_code=404)

            # Verifica se há produtos vinculados
            vinculados = conn.execute(
                'SELECT COUNT(*) as total FROM produtos WHERE almoxarifado_id = ?', (id,)
            ).fetchone()['total']
            if vinculados > 0:
                return response(
                    False,
                    message='Não é possível excluir almoxarifado com produtos vinculados.',
                    status_code=409
                )

            # Verifica se há estoque registrado neste almoxarifado
            estoque = conn.execute(
                'SELECT COUNT(*) as total FROM estoque WHERE almoxarifado_id = ?', (id,)
            ).fetchone()['total']
            if estoque > 0:
                return response(
                    False,
                    message='Não é possível excluir almoxarifado com estoque registrado. Transfira ou remova o estoque primeiro.',
                    status_code=409
                )

            conn.execute('DELETE FROM almoxarifados WHERE id = ?', (id,))
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'excluir', 'almoxarifados',
                              f'Excluiu almoxarifado "{row["nome"]}" (ID {id})')

        return response(True, message='Almoxarifado excluído com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


# ============================================================
# COLABORADORES
# ============================================================

@app.route('/colaboradores/<int:id>/codigo-barras', methods=['GET'])
def gerar_codigo_barras_colaborador(id):
    try:
        with get_connection() as conn:
            colab = conn.execute(
                'SELECT matricula, nome FROM colaboradores WHERE id = ?',
                (id,)
            ).fetchone()
            if not colab:
                return response(False, message='Colaborador não encontrado.', status_code=404)

            codigo = colab['matricula']
            if not codigo:
                return response(False, message='Colaborador sem matrícula.', status_code=400)

            CODE128 = barcode.get_barcode_class('code128')
            codigo_barras = CODE128(codigo, writer=ImageWriter())
            buf = io.BytesIO()
            codigo_barras.write(buf)
            buf.seek(0)
            return send_file(buf, mimetype='image/png')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/colaboradores/<int:id>/etiqueta', methods=['GET'])
def etiqueta_colaborador(id):
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401
    try:
        with get_connection() as conn:
            colab = conn.execute(
                'SELECT nome, matricula, setor, cargo FROM colaboradores WHERE id = ?',
                (id,)
            ).fetchone()
            if not colab:
                return '<h2>Colaborador não encontrado</h2>', 404

            nome = colab['nome'] or 'Sem nome'
            matricula = colab['matricula'] or '000000'
            setor = colab['setor'] or ''
            cargo = colab['cargo'] or ''

            html = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Etiqueta - {escape(nome)}</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:Arial,sans-serif; display:flex; flex-direction:column; align-items:center; justify-content:center; min-height:100vh; padding:20px; }}
    .etiqueta {{ width:280px; padding:16px; border:2px solid #333; border-radius:8px; text-align:center; background:#fff; }}
    .etiqueta img {{ width:220px; height:auto; margin-bottom:8px; }}
    .etiqueta .nome {{ font-size:14px; font-weight:bold; margin-bottom:4px; }}
    .etiqueta .info {{ font-size:11px; color:#555; }}
    .etiqueta .cargo {{ font-size:11px; color:#2563eb; margin-top:2px; }}
    .btn-imprimir {{ margin-top:20px; padding:12px 24px; font-size:16px; cursor:pointer; background:#2563eb; color:#fff; border:none; border-radius:6px; }}
    @media print {{ .btn-imprimir {{ display:none; }} .etiqueta {{ border:1px solid #000; }} }}
</style></head>
<body>
    <button class="btn-imprimir" onclick="window.print()">🖨️ Imprimir Etiqueta</button>
    <div class="etiqueta">
        <img src="/colaboradores/{id}/codigo-barras" alt="Matrícula">
        <div class="nome">{escape(nome)}</div>
        <div class="info">Mat: {escape(matricula)}</div>
        <div class="info">{escape(setor)}</div>
        <div class="cargo">{escape(cargo)}</div>
    </div>
</body></html>'''
            return html
    except Exception as e:
        return f'<h2>Erro: {str(e)}</h2>', 500

@app.route('/colaboradores/etiquetas/lote', methods=['GET'])
def etiquetas_colaboradores_lote():
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401

    ids = request.args.get('ids', '')
    if not ids:
        return '<h2>Nenhum colaborador selecionado</h2>', 400

    lista_ids = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not lista_ids:
        return '<h2>IDs inválidos</h2>', 400

    placeholders = ','.join(['?'] * len(lista_ids))
    with get_connection() as conn:
        colaboradores = conn.execute(
            f'SELECT id, nome, matricula, setor, cargo FROM colaboradores WHERE id IN ({placeholders})',
            lista_ids
        ).fetchall()

    if not colaboradores:
        return '<h2>Nenhum colaborador encontrado</h2>', 404

    etiquetas_html = ''
    for c in colaboradores:
        nome = c['nome'] or 'Sem nome'
        matricula = c['matricula'] or '000000'
        setor = c['setor'] or ''
        cargo = c['cargo'] or ''
        etiquetas_html += f'''
        <div class="etiqueta">
            <img src="/colaboradores/{c['id']}/codigo-barras" alt="Matrícula">
            <div class="nome">{escape(nome)}</div>
            <div class="info">Mat: {escape(matricula)}</div>
            <div class="info">{escape(setor)}</div>
            <div class="cargo">{escape(cargo)}</div>
        </div>'''

    html = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Etiquetas Colaboradores - Lote</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:Arial,Helvetica,sans-serif; background:#ccc; }}
    .page {{ width:210mm; min-height:297mm; margin:0 auto; background:#fff; padding:8mm 8mm 0 8mm; }}
    .grid {{ display:grid; grid-template-columns:repeat(6, 31mm); grid-template-rows:repeat(16, 17mm); gap:0; justify-content:space-between; }}
    .etiqueta {{ width:31mm; height:17mm; border:0.3mm dashed #999; display:flex; flex-direction:column; align-items:center; justify-content:center; padding:1mm; overflow:hidden; }}
    .etiqueta img {{ max-width:26mm; max-height:7mm; margin-bottom:0.5mm; }}
    .etiqueta .nome {{ font-size:5.5px; font-weight:bold; text-align:center; line-height:1.1; }}
    .etiqueta .info {{ font-size:5px; text-align:center; line-height:1.1; color:#333; }}
    .etiqueta .cargo {{ font-size:4.5px; text-align:center; line-height:1.1; color:#2563eb; }}
    .no-print {{ text-align:center; padding:10px 0; }}
    .no-print button {{ padding:10px 24px; font-size:14px; cursor:pointer; background:#2563eb; color:#fff; border:none; border-radius:6px; margin:0 4px; }}
    @media print {{ @page {{ size:A4; margin:0; }} body {{ background:#fff; }} .page {{ width:100%; padding:8mm 8mm 0 8mm; }} .no-print {{ display:none; }} .etiqueta {{ border:0.2mm solid #000; }} }}
</style></head>
<body>
    <div class="no-print">
        <button onclick="window.print()">🖨️ Imprimir ({len(colaboradores)} etiquetas)</button>
        <button onclick="window.close()">✕ Fechar</button>
        <p style="margin-top:8px;font-size:12px;color:#555;">Formato A4 · 6 colunas × 16 linhas · {len(colaboradores)} etiqueta(s)</p>
    </div>
    <div class="page"><div class="grid">{etiquetas_html}</div></div>
</body></html>'''
    return html

@app.route('/colaboradores', methods=['GET'])
@jwt_required()
def listar_colaboradores():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                'SELECT * FROM colaboradores ORDER BY nome'
            ).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/colaboradores', methods=['POST'])
@jwt_required()
def criar_colaborador():
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()

    if not nome:
        return response(False, message='Nome é obrigatório.', status_code=400)

    matricula = data.get('matricula', '').strip()

    # Verificar se matrícula já existe (se foi informada)
    if matricula:
        with get_connection() as conn:
            existente = conn.execute(
                'SELECT id FROM colaboradores WHERE matricula = ?', (matricula,)
            ).fetchone()
            if existente:
                return response(False, message='Já existe um colaborador com esta matrícula.', status_code=400)

    valores = {
        'nome': nome,
        'matricula': matricula,
        'codigo_barras': matricula,
        'email': data.get('email'),
        'setor': data.get('setor'),
        'telefone': data.get('telefone'),
        'cargo': data.get('cargo'),
        'status': data.get('status'),
        'created_at': now_iso(),
        'updated_at': now_iso()
    }

    colunas = ', '.join(valores.keys())
    placeholders = ', '.join(['?' for _ in valores])

    try:
        with get_connection() as conn:
            conn.execute(
                f'INSERT INTO colaboradores ({colunas}) VALUES ({placeholders})',
                tuple(valores.values())
            )
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'criar', 'colaboradores',
                              f'Criou colaborador "{nome}"')

        return response(True, message='Colaborador criado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/colaboradores/<int:id>', methods=['PUT'])
@jwt_required()
def atualizar_colaborador(id):
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()

    if not nome:
        return response(False, message='Nome é obrigatório.', status_code=400)

    matricula = data.get('matricula', '').strip()

    # Verificar se matrícula já existe (ignorando o próprio ID)
    if matricula:
        with get_connection() as conn:
            existente = conn.execute(
                'SELECT id FROM colaboradores WHERE matricula = ? AND id != ?', (matricula, id)
            ).fetchone()
            if existente:
                return response(False, message='Já existe outro colaborador com esta matrícula.', status_code=400)

    valores = {
        'nome': nome,
        'matricula': matricula,
        'email': data.get('email'),
        'setor': data.get('setor'),
        'telefone': data.get('telefone'),
        'cargo': data.get('cargo'),
        'status': data.get('status'),
        'updated_at': now_iso()
    }

    set_clause = ', '.join([f'{campo} = ?' for campo in valores.keys()])
    params = tuple(valores.values()) + (id,)

    try:
        with get_connection() as conn:
            cur = conn.execute(
                f'UPDATE colaboradores SET {set_clause} WHERE id = ?', params
            )
            conn.commit()
            if cur.rowcount == 0:
                return response(False, message='Colaborador não encontrado.', status_code=404)
        return response(True, message='Colaborador atualizado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/colaboradores/<int:id>', methods=['DELETE'])
@jwt_required()
def excluir_colaborador(id):
    try:
        with get_connection() as conn:
            # Primeiro busca o nome para o log
            row = conn.execute('SELECT nome FROM colaboradores WHERE id = ?', (id,)).fetchone()
            if not row:
                return response(False, message='Colaborador não encontrado.', status_code=404)

            conn.execute('DELETE FROM colaboradores WHERE id = ?', (id,))
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'excluir', 'colaboradores',
                              f'Excluiu colaborador "{row["nome"]}" (ID {id})')

        return response(True, message='Colaborador excluído com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

            # Exportar Importar Colaboradores

@app.route('/colaboradores/exportar', methods=['GET'])
@jwt_required()
def exportar_colaboradores():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                'SELECT nome, matricula, telefone, email, setor, cargo, status FROM colaboradores ORDER BY nome'
            ).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Colaboradores"

        # Cabeçalho
        cabecalhos = ['Nome', 'Matrícula', 'Telefone', 'E-mail', 'Setor', 'Cargo/Função', 'Status']
        ws.append(cabecalhos)

        # Negrito nos cabeçalhos
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Dados
        for row in rows:
            ws.append([row['nome'], row['matricula'], row['telefone'], row['email'],
                       row['setor'], row['cargo'], row['status']])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='colaboradores.xlsx'
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/colaboradores/importar', methods=['POST'])
@jwt_required()
def importar_colaboradores():
    if 'file' not in request.files:
        return response(False, message='Nenhum arquivo enviado.', status_code=400)

    file = request.files['file']
    if file.filename == '':
        return response(False, message='Arquivo sem nome.', status_code=400)

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        importados = 0
        ignorados = 0
        erros = []

        with get_connection() as conn:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                nome = str(row[0]).strip() if row[0] else ''
                matricula = str(row[1]).strip() if row[1] else ''
                telefone = str(row[2]).strip() if row[2] else ''
                email = str(row[3]).strip() if row[3] else ''
                setor = str(row[4]).strip() if row[4] else ''
                cargo = str(row[5]).strip() if row[5] else ''
                status = str(row[6]).strip() if row[6] else ''

                if not nome:
                    ignorados += 1
                    continue

                # Verificar se já existe pela matrícula
                if matricula:
                    existente = conn.execute(
                        'SELECT id FROM colaboradores WHERE matricula = ?', (matricula,)
                    ).fetchone()

                    if existente:
                        # Atualiza
                        conn.execute("""
                            UPDATE colaboradores SET nome=?, telefone=?, email=?, setor=?, cargo=?, status=?, updated_at=?
                            WHERE id=?
                        """, (nome, telefone, email, setor, cargo, status, now_iso(), existente['id']))
                        importados += 1
                        continue

                # Insere novo
                conn.execute("""
                    INSERT INTO colaboradores (nome, matricula, telefone, email, setor, cargo, status, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (nome, matricula, telefone, email, setor, cargo, status, now_iso(), now_iso()))
                importados += 1

            conn.commit()

            # Log
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'importar', 'colaboradores',
                              f'Importou {importados} colaboradores via Excel')

        return response(True, message=f'Importação concluída! {importados} registros processados.')
    except Exception as e:
        return response(False, message=f'Erro ao importar: {str(e)}', status_code=500)


# ============================================================
# UNIDADES
# ============================================================

@app.route('/unidades/<int:id>/codigo-barras', methods=['GET'])
def gerar_codigo_barras_unidade(id):
    try:
        with get_connection() as conn:
            unid = conn.execute(
                'SELECT tag, numero_serie FROM unidades WHERE id = ?',
                (id,)
            ).fetchone()
            if not unid:
                return response(False, message='Unidade não encontrada.', status_code=404)

            codigo = unid['tag'] or unid['numero_serie'] or 'SEMCODIGO'
            CODE128 = barcode.get_barcode_class('code128')
            codigo_barras = CODE128(codigo, writer=ImageWriter())
            buf = io.BytesIO()
            codigo_barras.write(buf)
            buf.seek(0)
            return send_file(buf, mimetype='image/png')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/<int:id>/etiqueta', methods=['GET'])
def etiqueta_unidade(id):
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401
    try:
        with get_connection() as conn:
            unid = conn.execute('''
                SELECT u.tag, u.numero_serie, u.status, p.nome as produto_nome
                FROM unidades u
                LEFT JOIN produtos p ON u.produto_id = p.id
                WHERE u.id = ?
            ''', (id,)).fetchone()
            if not unid:
                return '<h2>Unidade não encontrada</h2>', 404

            tag = unid['tag'] or 'SEMTAG'
            serie = unid['numero_serie'] or ''
            produto = unid['produto_nome'] or 'Sem produto'
            status = unid['status'] or ''

            html = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Etiqueta - {escape(tag)}</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:Arial,sans-serif; display:flex; flex-direction:column; align-items:center; justify-content:center; min-height:100vh; padding:20px; }}
    .etiqueta {{ width:280px; padding:16px; border:2px solid #333; border-radius:8px; text-align:center; background:#fff; }}
    .etiqueta img {{ width:220px; height:auto; margin-bottom:8px; }}
    .etiqueta .tag {{ font-size:16px; font-weight:bold; margin-bottom:2px; }}
    .etiqueta .produto {{ font-size:13px; margin-bottom:2px; }}
    .etiqueta .info {{ font-size:11px; color:#555; }}
    .etiqueta .status {{ font-size:11px; font-weight:bold; margin-top:2px; color:#2563eb; }}
    .btn-imprimir {{ margin-top:20px; padding:12px 24px; font-size:16px; cursor:pointer; background:#2563eb; color:#fff; border:none; border-radius:6px; }}
    @media print {{ .btn-imprimir {{ display:none; }} .etiqueta {{ border:1px solid #000; }} }}
</style></head>
<body>
    <button class="btn-imprimir" onclick="window.print()">🖨️ Imprimir Etiqueta</button>
    <div class="etiqueta">
        <img src="/unidades/{id}/codigo-barras" alt="Código">
        <div class="tag">{escape(tag)}</div>
        <div class="produto">{escape(produto)}</div>
        <div class="info">Série: {escape(serie)}</div>
        <div class="status">{escape(status)}</div>
    </div>
</body></html>'''
            return html
    except Exception as e:
        return f'<h2>Erro: {str(e)}</h2>', 500

@app.route('/unidades/etiquetas/lote', methods=['GET'])
def etiquetas_unidades_lote():
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401

    ids = request.args.get('ids', '')
    if not ids:
        return '<h2>Nenhuma unidade selecionada</h2>', 400

    lista_ids = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not lista_ids:
        return '<h2>IDs inválidos</h2>', 400

    placeholders = ','.join(['?'] * len(lista_ids))
    with get_connection() as conn:
        unidades = conn.execute(f'''
            SELECT u.id, u.tag, u.numero_serie, u.status, p.nome as produto_nome
            FROM unidades u
            LEFT JOIN produtos p ON u.produto_id = p.id
            WHERE u.id IN ({placeholders})
        ''', lista_ids).fetchall()

    if not unidades:
        return '<h2>Nenhuma unidade encontrada</h2>', 404

    etiquetas_html = ''
    for u in unidades:
        tag = u['tag'] or 'SEMTAG'
        serie = u['numero_serie'] or ''
        produto = u['produto_nome'] or 'Sem produto'
        status = u['status'] or ''
        etiquetas_html += f'''
        <div class="etiqueta">
            <img src="/unidades/{u['id']}/codigo-barras" alt="Código">
            <div class="tag">{escape(tag)}</div>
            <div class="produto">{escape(produto)}</div>
            <div class="info">Série: {escape(serie)}</div>
            <div class="status">{escape(status)}</div>
        </div>'''

    html = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Etiquetas Unidades - Lote</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:Arial,Helvetica,sans-serif; background:#ccc; }}
    .page {{ width:210mm; min-height:297mm; margin:0 auto; background:#fff; padding:8mm 8mm 0 8mm; }}
    .grid {{ display:grid; grid-template-columns:repeat(6, 31mm); grid-template-rows:repeat(16, 17mm); gap:0; justify-content:space-between; }}
    .etiqueta {{ width:31mm; height:17mm; border:0.3mm dashed #999; display:flex; flex-direction:column; align-items:center; justify-content:center; padding:0.8mm; overflow:hidden; }}
    .etiqueta img {{ max-width:26mm; max-height:6mm; margin-bottom:0.3mm; }}
    .etiqueta .tag {{ font-size:6px; font-weight:bold; text-align:center; line-height:1.1; }}
    .etiqueta .produto {{ font-size:5px; text-align:center; line-height:1.1; color:#333; }}
    .etiqueta .info {{ font-size:4.5px; text-align:center; line-height:1.1; color:#555; }}
    .etiqueta .status {{ font-size:4.5px; text-align:center; line-height:1.1; color:#2563eb; font-weight:bold; }}
    .no-print {{ text-align:center; padding:10px 0; }}
    .no-print button {{ padding:10px 24px; font-size:14px; cursor:pointer; background:#2563eb; color:#fff; border:none; border-radius:6px; margin:0 4px; }}
    @media print {{ @page {{ size:A4; margin:0; }} body {{ background:#fff; }} .page {{ width:100%; padding:8mm 8mm 0 8mm; }} .no-print {{ display:none; }} .etiqueta {{ border:0.2mm solid #000; }} }}
</style></head>
<body>
    <div class="no-print">
        <button onclick="window.print()">🖨️ Imprimir ({len(unidades)} etiquetas)</button>
        <button onclick="window.close()">✕ Fechar</button>
        <p style="margin-top:8px;font-size:12px;color:#555;">Formato A4 · 6 colunas × 16 linhas · {len(unidades)} etiqueta(s)</p>
    </div>
    <div class="page"><div class="grid">{etiquetas_html}</div></div>
</body></html>'''
    return html

@app.route('/unidades', methods=['GET'])
@jwt_required()
def listar_unidades():
    try:
        with get_connection() as conn:
            rows = conn.execute(
                '''SELECT u.*, p.nome as produto_nome, a.nome as almoxarifado_nome
                   FROM unidades u
                   LEFT JOIN produtos p ON u.produto_id = p.id
                   LEFT JOIN almoxarifados a ON u.almoxarifado_id = a.id
                   ORDER BY u.tag'''
            ).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/unidades', methods=['POST'])
@jwt_required()
def criar_unidade():
    data = request.get_json() or {}
    produto_id = data.get('produto_id')
    if not produto_id:
        return response(False, message='produto_id é obrigatório.', status_code=400)

    # Converte boolean para int (SQLite armazena 0/1)
    requer_calibracao_raw = data.get('requer_calibracao', False)
    requer_calibracao = True if requer_calibracao_raw else False

    valores = {
        'produto_id': produto_id,
        'tag': data.get('tag'),
        'numero_serie': data.get('numero_serie'),
        'status': data.get('status') or 'disponivel',
        'localizacao': data.get('localizacao'),
        'data_aquisicao': data.get('data_aquisicao'),
        'almoxarifado_id': data.get('almoxarifado_id'),
        'requer_calibracao': requer_calibracao,
        'data_validade_calibracao': data.get('data_validade_calibracao'),
        'numero_certificado': data.get('numero_certificado'),
        'data_ultima_manutencao': data.get('data_ultima_manutencao'),
        'status_manutencao': data.get('status_manutencao') or 'disponivel',
        'observacao': data.get('observacao'),
        'created_at': now_iso(),
        'updated_at': now_iso()
    }
    colunas = ', '.join(valores.keys())
    placeholders = ', '.join(['?' for _ in valores])
    try:
        with get_connection() as conn:
            conn.execute(
                f'INSERT INTO unidades ({colunas}) VALUES ({placeholders})',
                tuple(valores.values())
            )
            # Incrementa +1 no estoque do produto vinculado
            almoxarifado_id = valores.get('almoxarifado_id')
            if almoxarifado_id:
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                       VALUES (?, ?, 1)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + 1''',
                    (produto_id, almoxarifado_id)
                )
                # Cria movimentação de entrada para rastreabilidade
                tag = valores.get('tag') or 'sem tag'
                conn.execute('''
                    INSERT INTO movimentacoes
                        (produto_id, almoxarifado_id, tipo, quantidade,
                         valor_unitario, documento, observacao)
                    VALUES (?, ?, 'entrada', 1, 0, 'Cadastro de Unidade',
                           ?)
                ''', (produto_id, almoxarifado_id,
                      f'Entrada automatica - Cadastro de unidade TAG {tag}'))
            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                tag = valores.get('tag') or 'sem tag'
                registrar_log(user['id'], current_user, 'criar', 'unidades',
                              f'Cadastrou unidade "{tag}" (produto_id {produto_id})')
        return response(True, message='Unidade criada com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/unidades/<int:id>', methods=['PUT'])
@jwt_required()
def atualizar_unidade(id):
    data = request.get_json() or {}
    produto_id = data.get('produto_id')
    if not produto_id:
        return response(False, message='produto_id é obrigatório.', status_code=400)

    # Converte boolean para int (SQLite armazena 0/1)
    requer_calibracao_raw = data.get('requer_calibracao', False)
    requer_calibracao = True if requer_calibracao_raw else False

    valores = {
        'produto_id': produto_id,
        'tag': data.get('tag'),
        'numero_serie': data.get('numero_serie'),
        'status': data.get('status'),
        'localizacao': data.get('localizacao'),
        'data_aquisicao': data.get('data_aquisicao'),
        'almoxarifado_id': data.get('almoxarifado_id'),
        'requer_calibracao': requer_calibracao,
        'data_validade_calibracao': data.get('data_validade_calibracao'),
        'numero_certificado': data.get('numero_certificado'),
        'data_ultima_manutencao': data.get('data_ultima_manutencao'),
        'status_manutencao': data.get('status_manutencao') or 'disponivel',
        'observacao': data.get('observacao'),
        'updated_at': now_iso()
    }
    set_clause = ', '.join([f'{campo} = ?' for campo in valores.keys()])
    params = tuple(valores.values()) + (id,)
    try:
        with get_connection() as conn:
            cur = conn.execute(
                f'UPDATE unidades SET {set_clause} WHERE id = ?', params
            )
            conn.commit()
            if cur.rowcount == 0:
                return response(False, message='Unidade não encontrada.', status_code=404)
        return response(True, message='Unidade atualizada com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/unidades/<int:id>', methods=['DELETE'])
@jwt_required()
def excluir_unidade(id):
    try:
        with get_connection() as conn:
            # Primeiro busca os dados da unidade para o log
            row = conn.execute('''
                SELECT u.tag, p.nome as produto_nome
                FROM unidades u
                LEFT JOIN produtos p ON p.id = u.produto_id
                WHERE u.id = ?
            ''', (id,)).fetchone()
            if not row:
                return response(False, message='Unidade não encontrada.', status_code=404)

            conn.execute('DELETE FROM unidades WHERE id = ?', (id,))
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                tag = row['tag'] or 'sem tag'
                produto = row['produto_nome'] or 'produto desconhecido'
                registrar_log(user['id'], current_user, 'excluir', 'unidades',
                              f'Excluiu unidade "{tag}" (ID {id}) do produto "{produto}"')

        return response(True, message='Unidade excluída com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/<int:id>/manutencao', methods=['POST'])
@jwt_required()
def iniciar_manutencao_unidade(id):
    data = request.get_json() or {}
    descricao = data.get('descricao')
    if not descricao:
        return response(False, message='descricao é obrigatória.', status_code=400)
    fornecedor = data.get('fornecedor')
    try:
        with get_connection() as conn:
            row = conn.execute(
                'SELECT tag, produto_id, almoxarifado_id FROM unidades WHERE id = ?', (id,)
            ).fetchone()
            if not row:
                return response(False, message='Unidade não encontrada.', status_code=404)

            produto_id = row['produto_id']
            almoxarifado_origem_id = row['almoxarifado_id']
            tag = row['tag'] or 'sem tag'

            # Insere registro de manutenção com almoxarifado de origem
            conn.execute(
                '''INSERT INTO manutencoes_unidades
                   (unidade_id, descricao, fornecedor, data_envio, status, almoxarifado_origem_id, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?)''',
                (id, descricao, fornecedor, now_iso(), 'em_manutencao', almoxarifado_origem_id, now_iso())
            )

            # Atualiza status da unidade
            conn.execute(
                'UPDATE unidades SET status_manutencao = ?, updated_at = ? WHERE id = ?',
                ('em_manutencao', now_iso(), id)
            )

            # Decrementa estoque da origem
            if almoxarifado_origem_id:
                conn.execute(
                    '''UPDATE estoque
                       SET quantidade = quantidade - 1, updated_at = ?
                       WHERE produto_id = ? AND almoxarifado_id = ?''',
                    (now_iso(), produto_id, almoxarifado_origem_id)
                )

                # Cria movimentação de saída
                conn.execute('''
                    INSERT INTO movimentacoes
                        (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario,
                         documento, observacao)
                    VALUES (?, ?, 'saida', 1, 0, 'Manutenção', ?)
                ''', (produto_id, almoxarifado_origem_id,
                      f'Saída - Envio para manutenção - TAG {tag}'))

            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'manutencao', 'unidades',
                              f'Enviou unidade "{tag}" (ID {id}) para manutenção')
        return response(True, message='Unidade enviada para manutenção com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/<int:id>/manutencao/retorno', methods=['POST'])
@jwt_required()
def retornar_manutencao_unidade(id):
    data = request.get_json() or {}
    custo = data.get('custo')
    destino_id = data.get('destino_id')
    try:
        with get_connection() as conn:
            row = conn.execute(
                'SELECT tag, produto_id, almoxarifado_id FROM unidades WHERE id = ?', (id,)
            ).fetchone()
            if not row:
                return response(False, message='Unidade não encontrada.', status_code=404)

            produto_id = row['produto_id']
            tag = row['tag'] or 'sem tag'

            # Se não informou destino, usa o almoxarifado atual da unidade
            if not destino_id:
                destino_id = row['almoxarifado_id']

            # Busca manutenção em aberto
            manut = conn.execute(
                '''SELECT id, almoxarifado_origem_id FROM manutencoes_unidades
                   WHERE unidade_id = ? AND status = 'em_manutencao'
                   ORDER BY created_at DESC LIMIT 1''',
                (id,)
            ).fetchone()
            if not manut:
                return response(False, message='Nenhuma manutenção em aberto para esta unidade.', status_code=404)

            # Atualiza manutenção como concluída
            conn.execute(
                '''UPDATE manutencoes_unidades
                   SET status = 'concluida', data_retorno = ?, custo = ?
                   WHERE id = ?''',
                (now_iso(), custo, manut['id'])
            )

            # Atualiza unidade: status + almoxarifado de destino
            conn.execute(
                '''UPDATE unidades
                   SET status_manutencao = 'disponivel',
                       almoxarifado_id = ?,
                       data_ultima_manutencao = ?, updated_at = ?
                   WHERE id = ?''',
                (destino_id, now_iso(), now_iso(), id)
            )

            # Incrementa estoque no destino
            conn.execute(
                '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at)
                   VALUES (?, ?, 1, COALESCE((SELECT estoque_minimo FROM produtos WHERE id = ?), 0), ?)
                   ON CONFLICT(produto_id, almoxarifado_id)
                   DO UPDATE SET quantidade = estoque.quantidade + 1, updated_at = excluded.updated_at''',
                (produto_id, destino_id, produto_id, now_iso())
            )

            # Movimentação de entrada
            conn.execute('''
                INSERT INTO movimentacoes
                    (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario,
                     documento, observacao)
                VALUES (?, ?, 'entrada', 1, ?, 'Retorno Manutenção', ?)
            ''', (produto_id, destino_id, custo or 0,
                  f'Entrada - Retorno de manutenção - TAG {tag}'))

            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'manutencao_retorno', 'unidades',
                              f'Retornou unidade "{tag}" (ID {id}) da manutenção')
        return response(True, message='Unidade retornada da manutenção com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/manutencao/retorno/lote', methods=['POST'])
@jwt_required()
def retornar_manutencao_lote():
    data = request.get_json() or {}
    unidades_ids = data.get('unidades_ids', [])
    destino_id = data.get('destino_id')
    custo = data.get('custo')

    if not unidades_ids:
        return response(False, message='Selecione ao menos uma unidade.', status_code=400)
    if not destino_id:
        return response(False, message='destino_id é obrigatório.', status_code=400)

    try:
        with get_connection() as conn:
            destino = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (destino_id,)).fetchone()
            if not destino:
                return response(False, message='Almoxarifado de destino não encontrado.', status_code=404)

            doc_retorno = f"Retorno Manut. Lote {now_iso()[:19]}"
            retornadas = 0
            erros = []

            for uid in unidades_ids:
                row = conn.execute(
                    'SELECT tag, produto_id FROM unidades WHERE id = ?', (uid,)
                ).fetchone()
                if not row:
                    erros.append(f'Unidade ID {uid} não encontrada.')
                    continue

                manut = conn.execute(
                    '''SELECT id FROM manutencoes_unidades
                       WHERE unidade_id = ? AND status = 'em_manutencao'
                       ORDER BY created_at DESC LIMIT 1''',
                    (uid,)
                ).fetchone()
                if not manut:
                    erros.append(f'Unidade TAG {row["tag"]} não tem manutenção em aberto.')
                    continue

                produto_id = row['produto_id']
                tag = row['tag'] or 'sem tag'

                # Fecha manutenção
                conn.execute(
                    '''UPDATE manutencoes_unidades
                       SET status = 'concluida', data_retorno = ?, custo = ?
                       WHERE id = ?''',
                    (now_iso(), custo, manut['id'])
                )

                # Atualiza unidade: status + almoxarifado destino
                conn.execute(
                    '''UPDATE unidades
                       SET status_manutencao = 'disponivel',
                           almoxarifado_id = ?,
                           data_ultima_manutencao = ?, updated_at = ?
                       WHERE id = ?''',
                    (destino_id, now_iso(), now_iso(), uid)
                )

                # Incrementa estoque no destino
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at)
                       VALUES (?, ?, 1, COALESCE((SELECT estoque_minimo FROM produtos WHERE id = ?), 0), ?)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + 1, updated_at = excluded.updated_at''',
                    (produto_id, destino_id, produto_id, now_iso())
                )

                # Movimentação de entrada
                conn.execute('''
                    INSERT INTO movimentacoes
                        (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario,
                         documento, observacao)
                    VALUES (?, ?, 'entrada', 1, ?, ?, ?)
                ''', (produto_id, destino_id, custo or 0, doc_retorno,
                      f'Entrada - Retorno manutenção em lote - TAG {tag}'))

                retornadas += 1

            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'manutencao_retorno_lote', 'unidades',
                              f'Retornou {retornadas} unidade(s) da manutenção em lote para "{destino["nome"]}"')

            msg = f'{retornadas} unidade(s) retornada(s) da manutenção.'
            if erros:
                msg += f' {len(erros)} erro(s): ' + '; '.join(erros)
            return response(True, message=msg)
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/exportar', methods=['GET'])
@jwt_required()
def exportar_unidades():
    try:
        with get_connection() as conn:
            rows = conn.execute('''
                SELECT u.tag, u.numero_serie, p.nome AS produto_nome, u.status,
                       u.localizacao, a.nome AS almoxarifado_nome
                FROM unidades u
                LEFT JOIN produtos p ON u.produto_id = p.id
                LEFT JOIN almoxarifados a ON u.almoxarifado_id = a.id
                ORDER BY u.tag
            ''').fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unidades"
        cabecalhos = ['TAG', 'Nº Série', 'Produto', 'Status', 'Localização', 'Almoxarifado']
        ws.append(cabecalhos)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for row in rows:
            ws.append([
                row['tag'], row['numero_serie'], row['produto_nome'],
                row['status'], row['localizacao'], row['almoxarifado_nome']
            ])
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='unidades.xlsx'
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/transferir', methods=['POST'])
@jwt_required()
def transferir_unidades_lote():
    data = request.get_json() or {}
    unidades_ids = data.get('unidades_ids', [])
    destino_id = data.get('destino_id')

    if not unidades_ids:
        return response(False, message='Selecione ao menos uma unidade.', status_code=400)
    if not destino_id:
        return response(False, message='destino_id é obrigatório.', status_code=400)

    try:
        with get_connection() as conn:
            destino = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (destino_id,)).fetchone()
            if not destino:
                return response(False, message='Almoxarifado de destino não encontrado.', status_code=404)

            transferidas = 0
            erros = []
            doc_transf = f"Transferência Lote {now_iso()[:19]}"

            for uid in unidades_ids:
                row = conn.execute(
                    'SELECT tag, produto_id, almoxarifado_id FROM unidades WHERE id = ?', (uid,)
                ).fetchone()
                if not row:
                    erros.append(f'Unidade ID {uid} não encontrada.')
                    continue
                if row['almoxarifado_id'] == destino_id:
                    erros.append(f'Unidade TAG {row["tag"]} já está no destino.')
                    continue

                produto_id = row['produto_id']
                origem_id = row['almoxarifado_id']
                tag = row['tag'] or 'sem tag'

                # Atualiza almoxarifado e localização da unidade
                conn.execute(
                    'UPDATE unidades SET almoxarifado_id = ?, localizacao = ?, updated_at = ? WHERE id = ?',
                    (destino_id, destino['nome'], now_iso(), uid)
                )

                # Decrementa estoque da origem
                if origem_id:
                    conn.execute(
                        '''UPDATE estoque SET quantidade = quantidade - 1, updated_at = ?
                           WHERE produto_id = ? AND almoxarifado_id = ?''',
                        (now_iso(), produto_id, origem_id)
                    )
                    conn.execute('''
                        INSERT INTO movimentacoes (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario, documento, observacao)
                        VALUES (?, ?, 'saida', 1, 0, ?, ?)
                    ''', (produto_id, origem_id, doc_transf, f'Saída - Transferência em lote - TAG {tag}'))

                # Incrementa estoque do destino
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at)
                       VALUES (?, ?, 1, COALESCE((SELECT estoque_minimo FROM produtos WHERE id = ?), 0), ?)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + 1, updated_at = excluded.updated_at''',
                    (produto_id, destino_id, produto_id, now_iso())
                )
                conn.execute('''
                    INSERT INTO movimentacoes (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario, documento, observacao)
                    VALUES (?, ?, 'entrada', 1, 0, ?, ?)
                ''', (produto_id, destino_id, doc_transf, f'Entrada - Transferência em lote - TAG {tag}'))

                transferidas += 1

            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'transferencia_lote', 'unidades',
                              f'Transferiu {transferidas} unidade(s) para "{destino["nome"]}"')

            msg = f'{transferidas} unidade(s) transferida(s).'
            if erros:
                msg += f' {len(erros)} erro(s): ' + '; '.join(erros)
            return response(True, message=msg)
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/importar', methods=['POST'])
@jwt_required()
def importar_unidades():
    if 'file' not in request.files:
        return response(False, message='Nenhum arquivo enviado.', status_code=400)
    file = request.files['file']
    if file.filename == '':
        return response(False, message='Arquivo sem nome.', status_code=400)

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        importados = 0
        erros = []

        with get_connection() as conn:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                tag = str(row[0]).strip() if row[0] else ''
                numero_serie = str(row[1]).strip() if row[1] else ''
                produto_nome = str(row[2]).strip() if row[2] else ''
                status = str(row[3]).strip() if row[3] else 'disponivel'
                localizacao = str(row[4]).strip() if row[4] else ''
                almoxarifado_nome = str(row[5]).strip() if row[5] else ''

                if not tag:
                    erros.append(f'Linha {i}: TAG é obrigatória')
                    continue

                # Busca o produto pelo nome
                produto = conn.execute(
                    'SELECT id FROM produtos WHERE nome = ?', (produto_nome,)
                ).fetchone()
                if not produto:
                    erros.append(f'Linha {i}: produto "{produto_nome}" não encontrado')
                    continue

                # Busca o almoxarifado pelo nome
                almoxarifado_id = None
                if almoxarifado_nome:
                    almox = conn.execute(
                        'SELECT id FROM almoxarifados WHERE nome = ?', (almoxarifado_nome,)
                    ).fetchone()
                    if almox:
                        almoxarifado_id = almox['id']

                # Verifica se já existe pela TAG
                existente = conn.execute(
                    'SELECT id FROM unidades WHERE tag = ?', (tag,)
                ).fetchone()

                if existente:
                    conn.execute("""
                        UPDATE unidades SET numero_serie=?, produto_id=?, status=?,
                            localizacao=?, almoxarifado_id=?, updated_at=?
                        WHERE id=?
                    """, (numero_serie, produto['id'], status, localizacao,
                          almoxarifado_id, now_iso(), existente['id']))
                else:
                    conn.execute("""
                        INSERT INTO unidades (produto_id, tag, numero_serie, status,
                            localizacao, almoxarifado_id, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (produto['id'], tag, numero_serie, status, localizacao,
                          almoxarifado_id, now_iso(), now_iso()))
                importados += 1

            conn.commit()

        msg = f'{importados} unidade(s) processada(s).'
        if erros:
            msg += f' {len(erros)} erro(s): ' + '; '.join(erros[:5])
            if len(erros) > 5:
                msg += f' (e mais {len(erros)-5})'
        return response(True, message=msg)
    except Exception as e:
        return response(False, message=f'Erro ao importar: {str(e)}', status_code=500)

@app.route('/unidades/manutencao/lote', methods=['POST'])
@jwt_required()
def enviar_unidades_manutencao_lote():
    data = request.get_json() or {}
    unidades_ids = data.get('unidades_ids', [])
    descricao = data.get('descricao')
    fornecedor = data.get('fornecedor')

    if not unidades_ids:
        return response(False, message='Selecione ao menos uma unidade.', status_code=400)
    if not descricao:
        return response(False, message='descricao é obrigatória.', status_code=400)

    try:
        with get_connection() as conn:
            enviadas = 0
            erros = []
            doc_manut = f"Manut. Lote {now_iso()[:19]}"

            for uid in unidades_ids:
                row = conn.execute(
                    'SELECT tag, produto_id, almoxarifado_id, status_manutencao FROM unidades WHERE id = ?', (uid,)
                ).fetchone()
                if not row:
                    erros.append(f'Unidade ID {uid} não encontrada.')
                    continue
                if row['status_manutencao'] == 'em_manutencao':
                    erros.append(f'Unidade TAG {row["tag"]} já está em manutenção.')
                    continue

                produto_id = row['produto_id']
                almoxarifado_origem_id = row['almoxarifado_id']
                tag = row['tag'] or 'sem tag'

                # Insere registro de manutenção
                conn.execute(
                    '''INSERT INTO manutencoes_unidades
                       (unidade_id, descricao, fornecedor, data_envio, status,
                        almoxarifado_origem_id, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (uid, descricao, fornecedor, now_iso(), 'em_manutencao',
                     almoxarifado_origem_id, now_iso())
                )

                # Atualiza status da unidade
                conn.execute(
                    'UPDATE unidades SET status_manutencao = ?, updated_at = ? WHERE id = ?',
                    ('em_manutencao', now_iso(), uid)
                )

                # Decrementa estoque da origem
                if almoxarifado_origem_id:
                    conn.execute(
                        '''UPDATE estoque
                           SET quantidade = quantidade - 1, updated_at = ?
                           WHERE produto_id = ? AND almoxarifado_id = ?''',
                        (now_iso(), produto_id, almoxarifado_origem_id)
                    )
                    # Movimentação de saída
                    conn.execute('''
                        INSERT INTO movimentacoes
                            (produto_id, almoxarifado_id, tipo, quantidade, valor_unitario,
                             documento, observacao)
                        VALUES (?, ?, 'saida', 1, 0, ?, ?)
                    ''', (produto_id, almoxarifado_origem_id, doc_manut,
                          f'Saída - Manutenção em lote - TAG {tag}'))

                enviadas += 1

            conn.commit()
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'manutencao_lote', 'unidades',
                              f'Enviou {enviadas} unidade(s) para manutenção em lote')

            msg = f'{enviadas} unidade(s) enviada(s) para manutenção.'
            if erros:
                msg += f' {len(erros)} erro(s): ' + '; '.join(erros)
            return response(True, message=msg)
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/unidades/busca', methods=['GET'])
@jwt_required()
def buscar_unidade_por_tag():
    tag = request.args.get('tag', '').strip()
    if not tag:
        return response(False, message='Informe a TAG da unidade.', status_code=400)
    try:
        with get_connection() as conn:
            row = conn.execute("""
                SELECT u.id, u.tag, u.status, u.numero_serie, u.almoxarifado_id,
                       p.id as produto_id, p.nome as produto_nome, p.codigo_interno,
                       p.custo_medio
                FROM unidades u
                LEFT JOIN produtos p ON u.produto_id = p.id
                WHERE u.tag = ?
                LIMIT 1
            """, (tag,)).fetchone()
            if not row:
                return response(False, message='Unidade não encontrada', status_code=404)
            return response(True, data=dict(row))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/movimentacoes/emprestimo-rapido', methods=['POST'])
@jwt_required()
def emprestimo_rapido():
    data = request.get_json()
    if not data:
        return response(False, message='Dados inválidos.', status_code=400)
    
    colaborador_id = data.get('colaborador_id')
    almoxarifado_id = data.get('almoxarifado_id')
    itens = data.get('itens', [])
    
    if not colaborador_id or not itens:
        return response(False, message='Colaborador e itens são obrigatórios.', status_code=400)
    
    try:
        with get_connection() as conn:
            resultados = []
            for item in itens:
                unidade_id = item.get('unidade_id')
                produto_id = item.get('produto_id')
                
                if not unidade_id:
                    continue
                
                # Verifica se a unidade ainda está disponível
                unidade = conn.execute(
                    'SELECT id, tag, status FROM unidades WHERE id = ? AND status = ?',
                    (unidade_id, 'disponivel')
                ).fetchone()
                
                if not unidade:
                    return response(False, message=f'Unidade TAG {item.get("tag", unidade_id)} não está mais disponível.', status_code=409)
                
                # Cria o empréstimo cursor
                conn.execute("""
                    INSERT INTO movimentacoes (tipo, produto_id, equipamento_id, colaborador_id, almoxarifado_id, quantidade, observacao, natureza, data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, NOW() - INTERVAL '3 hours')
                """, ('saida', produto_id, unidade_id, colaborador_id, almoxarifado_id, 1, 'Empréstimo rápido via movimentação rápida', 'emprestimo'))

                # Cria registro na tabela emprestimos
                conn.execute("""
                    INSERT INTO emprestimos (unidade_id, colaborador_id, data_emprestimo, status, tipo)
                    VALUES (?, ?, NOW() - INTERVAL '3 hours', 'ativo', 'emprestimo')
                """, (unidade_id, colaborador_id))
                
                # Atualiza status da unidade
                conn.execute(
                    "UPDATE unidades SET status = 'emprestado', localizacao = (SELECT nome FROM colaboradores WHERE id = ?) WHERE id = ?",
                    (colaborador_id, unidade_id)
                )
                
                resultados.append({
                    'unidade_id': unidade_id,
                    'tag': item.get('tag', ''),
                    'produto_nome': item.get('produto_nome', ''),
                    'status': 'emprestado'
                })
            
            conn.commit()
            return response(True, message=f'{len(resultados)} unidade(s) emprestada(s) com sucesso!', data={'itens': resultados})
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/movimentacoes/devolucao-rapida', methods=['POST'])
@jwt_required()
def devolucao_rapida():
    data = request.get_json()
    if not data:
        return response(False, message='Dados inválidos.', status_code=400)
    
    colaborador_id = data.get('colaborador_id')
    almoxarifado_id = data.get('almoxarifado_id')
    itens = data.get('itens', [])
    
    if not itens:
        return response(False, message='Itens são obrigatórios.', status_code=400)
    
    try:
        with get_connection() as conn:
            for item in itens:
                unidade_id = item.get('unidade_id')
                produto_id = item.get('produto_id')
                
                if not unidade_id:
                    continue
                
                # Atualiza empréstimo ativo para devolvido
                conn.execute("""
                    UPDATE emprestimos 
                    SET status = 'devolvido', data_devolucao = NOW() - INTERVAL '3 hours'
                    WHERE unidade_id = ? AND status = 'ativo'
                """, (unidade_id,))
                
                # Atualiza status da unidade para disponível
                conn.execute(
                    "UPDATE unidades SET status = 'disponivel', localizacao = NULL WHERE id = ?",
                    (unidade_id,)
                )
                
                # Registra na movimentações como entrada
                conn.execute("""
                    INSERT INTO movimentacoes (tipo, produto_id, equipamento_id, colaborador_id, almoxarifado_id, quantidade, observacao, natureza, data)
                    VALUES ('entrada', ?, ?, ?, ?, 1, 'Devolução rápida via movimentação rápida', 'emprestimo', NOW() - INTERVAL '3 hours')
                """, (produto_id, unidade_id, colaborador_id, almoxarifado_id))
            
            conn.commit()
            return response(True, message=f'{len(itens)} unidade(s) devolvida(s) com sucesso!')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

# ============================================================
# ESTOQUE
# ============================================================

@app.route('/estoque/exportar', methods=['GET'])
@jwt_required()
def exportar_estoque():
    try:
        with get_connection() as conn:
            rows = conn.execute('''
                SELECT p.codigo_interno, p.nome AS produto_nome,
                       e.quantidade, a.nome AS almoxarifado_nome
                FROM estoque e
                LEFT JOIN produtos p ON e.produto_id = p.id
                LEFT JOIN almoxarifados a ON e.almoxarifado_id = a.id
                ORDER BY p.nome ASC
            ''').fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"
        # Headers compatíveis com o importador
        cabecalhos = ['codigo_interno', 'quantidade', 'almoxarifado']
        ws.append(cabecalhos)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for row in rows:
            ws.append([
                row['codigo_interno'],
                row['quantidade'],
                row['almoxarifado_nome'] or ''
            ])
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='estoque.xlsx'
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/estoque', methods=['GET'])
@jwt_required()
def listar_estoque():
    almoxarifado_id = request.args.get('almoxarifado_id')
    busca = request.args.get('busca')

    sql = '''
        SELECT e.id, e.produto_id, e.almoxarifado_id, e.quantidade, e.estoque_minimo,
               p.nome AS produto_nome, p.codigo_interno, p.categoria, p.custo_medio,
               p.estoque_minimo AS produto_estoque_minimo,
               e.quantidade * COALESCE(p.custo_medio, 0) AS valor_patrimonial,
               a.nome AS almoxarifado_nome,
                CASE WHEN EXISTS (
                   SELECT 1 FROM pedidos_itens pi
                   JOIN pedidos pd ON pd.id = pi.pedido_id
                   WHERE pi.produto_id = e.produto_id
                   AND pd.status IN ('aberto', 'em_compra')
               ) THEN 1 ELSE 0 END as possui_pedido_aberto
        FROM estoque e
        LEFT JOIN produtos p ON e.produto_id = p.id
        LEFT JOIN almoxarifados a ON e.almoxarifado_id = a.id
    '''

    conditions = []
    params = []

    if almoxarifado_id:
        conditions.append('e.almoxarifado_id = ?')
        params.append(almoxarifado_id)

    if busca:
        conditions.append('p.nome LIKE ?')
        params.append(f'%{busca}%')

    if conditions:
        sql += ' WHERE ' + ' AND '.join(conditions)

    sql += ' ORDER BY p.nome ASC'

    try:
        with get_connection() as conn:
            rows = conn.execute(sql, params).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/estoque', methods=['POST'])
@jwt_required()
def upsert_estoque():
    data = request.get_json() or {}
    produto_id = data.get('produto_id')
    almoxarifado_id = data.get('almoxarifado_id')
    quantidade = data.get('quantidade')
    estoque_minimo = data.get('estoque_minimo', 0)

    if not produto_id or not almoxarifado_id or quantidade is None:
        return response(
            False,
            message='produto_id, almoxarifado_id e quantidade são obrigatórios.',
            status_code=400
        )

    try:
        with get_connection() as conn:
            # Busca dados para o log (nome do produto e almoxarifado)
            produto = conn.execute('SELECT nome FROM produtos WHERE id = ?', (produto_id,)).fetchone()
            almox = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (almoxarifado_id,)).fetchone()
            nome_produto = produto['nome'] if produto else f'ID {produto_id}'
            nome_almox = almox['nome'] if almox else f'ID {almoxarifado_id}'

            # Busca quantidade anterior para comparar
            anterior = conn.execute(
                'SELECT quantidade, estoque_minimo FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                (produto_id, almoxarifado_id)
            ).fetchone()
            qtd_anterior = anterior['quantidade'] if anterior else 0

            # Define o estoque_minimo: valor enviado > 0 usa ele, senão copia do produto
            if anterior and anterior['estoque_minimo'] is not None and anterior['estoque_minimo'] > 0:
                minimo_anterior = anterior['estoque_minimo']
            else:
                minimo_anterior = 0
                
            if estoque_minimo and estoque_minimo > 0:
                minimo_final = estoque_minimo
            elif minimo_anterior > 0:
                minimo_final = minimo_anterior
            else:
                # Fallback: copia o estoque_minimo do produto
                prod_min = conn.execute(
                    'SELECT estoque_minimo FROM produtos WHERE id = ?', (produto_id,)
                ).fetchone()
                minimo_final = prod_min['estoque_minimo'] if prod_min and prod_min['estoque_minimo'] else 0

            conn.execute(
                '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(produto_id, almoxarifado_id)
                   DO UPDATE SET 
                       quantidade = excluded.quantidade,
                       estoque_minimo = CASE 
                           WHEN excluded.estoque_minimo > 0 THEN excluded.estoque_minimo 
                           ELSE estoque.estoque_minimo 
                       END,
                       updated_at = excluded.updated_at''',
                (produto_id, almoxarifado_id, quantidade, minimo_final, now_iso())
            )
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                diferenca = quantidade - qtd_anterior
                if qtd_anterior == 0:
                    descricao = f'Inseriu estoque de "{nome_produto}" em "{nome_almox}": {quantidade} unidade(s)'
                else:
                    descricao = f'Atualizou estoque de "{nome_produto}" em "{nome_almox}": {qtd_anterior} → {quantidade} ({diferenca:+d})'
                registrar_log(user['id'], current_user, 'editar', 'estoque', descricao)

        return response(True, message='Estoque atualizado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/estoque/transferir', methods=['POST'])
@jwt_required()
def transferir_estoque():
    data = request.get_json() or {}
    origem_id = data.get('origem_id')
    destino_id = data.get('destino_id')
    produto_id = data.get('produto_id')
    quantidade = data.get('quantidade')

    if not origem_id or not destino_id or not produto_id or quantidade is None:
        return response(
            False,
            message='origem_id, destino_id, produto_id e quantidade são obrigatórios.',
            status_code=400
        )

    try:
        with get_connection() as conn:
            # Busca dados para o log
            produto = conn.execute('SELECT nome FROM produtos WHERE id = ?', (produto_id,)).fetchone()
            origem = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (origem_id,)).fetchone()
            destino = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (destino_id,)).fetchone()
            nome_produto = produto['nome'] if produto else f'ID {produto_id}'
            nome_origem = origem['nome'] if origem else f'ID {origem_id}'
            nome_destino = destino['nome'] if destino else f'ID {destino_id}'

            saldo_origem = conn.execute(
                'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                (produto_id, origem_id)
            ).fetchone()

            saldo_atual = saldo_origem['quantidade'] if saldo_origem else 0
            if saldo_atual < quantidade:
                return response(False, message='Saldo insuficiente na origem.', status_code=409)

            conn.execute(
                '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at)
                   VALUES (?, ?, 0, COALESCE((SELECT estoque_minimo FROM produtos WHERE id = ?), 0), ?)
                   ON CONFLICT(produto_id, almoxarifado_id)
                   DO UPDATE SET quantidade = estoque.quantidade, updated_at = excluded.updated_at''',
                (produto_id, destino_id, produto_id, now_iso())
            )

            conn.execute(
                '''UPDATE estoque
                   SET quantidade = quantidade - ?, updated_at = ?
                   WHERE produto_id = ? AND almoxarifado_id = ?''',
                (quantidade, now_iso(), produto_id, origem_id)
            )

            conn.execute(
                '''UPDATE estoque
                   SET quantidade = quantidade + ?, updated_at = ?
                   WHERE produto_id = ? AND almoxarifado_id = ?''',
                (quantidade, now_iso(), produto_id, destino_id)
            )
            conn.commit()

            # Log da ação
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'transferir', 'estoque',
                              f'Transferiu {quantidade} unidade(s) de "{nome_produto}": "{nome_origem}" → "{nome_destino}"')

        return response(True, message='Transferência realizada com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/estoque/importar', methods=['POST'])
@jwt_required()
def importar_estoque():
    if 'file' not in request.files:
        return response(False, message='Nenhum arquivo enviado.', status_code=400)

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return response(False, message='Formato não suportado. Use .xlsx, .xls ou .csv.', status_code=400)

    import pandas as pd
    import io

    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(file.read().decode('utf-8-sig')))
        else:
            df = pd.read_excel(io.BytesIO(file.read()))
    except Exception as e:
        return response(False, message='Erro ao ler o arquivo: ' + str(e), status_code=400)

    # Normaliza nomes das colunas
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

    colunas_necessarias = ['codigo_interno', 'quantidade']
    if not all(c in df.columns for c in colunas_necessarias):
        return response(False, message=f'Colunas necessárias: codigo_interno, quantidade. Colunas encontradas: {", ".join(df.columns)}', status_code=400)

    importados = 0
    erros = []

    with get_connection() as conn:
        for idx, row in df.iterrows():
            codigo = str(row['codigo_interno']).strip()
            quantidade = row['quantidade']

            try:
                quantidade = int(quantidade)
            except (ValueError, TypeError):
                erros.append(f'Linha {idx+2}: quantidade inválida "{row["quantidade"]}"')
                continue

            # Busca o produto pelo código interno
            produto = conn.execute(
                'SELECT id FROM produtos WHERE codigo_interno = ?', (codigo,)
            ).fetchone()

            if not produto:
                erros.append(f'Linha {idx+2}: produto "{codigo}" não encontrado')
                continue

            # Se tiver coluna almoxarifado, usa; senão, usa o primeiro almoxarifado
            almoxarifado_id = None
            if 'almoxarifado' in df.columns:
                nome_almox = str(row['almoxarifado']).strip()
                almox = conn.execute(
                    'SELECT id FROM almoxarifados WHERE nome = ?', (nome_almox,)
                ).fetchone()
                if almox:
                    almoxarifado_id = almox['id']
                else:
                    erros.append(f'Linha {idx+2}: almoxarifado "{nome_almox}" não encontrado')
                    continue
            else:
                # Pega o primeiro almoxarifado cadastrado
                almox = conn.execute('SELECT id FROM almoxarifados LIMIT 1').fetchone()
                if not almox:
                    erros.append('Nenhum almoxarifado cadastrado')
                    break
                almoxarifado_id = almox['id']

            conn.execute(
                '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                   VALUES (?, ?, ?)
                   ON CONFLICT(produto_id, almoxarifado_id)
                   DO UPDATE SET quantidade = excluded.quantidade''',
                (produto['id'], almoxarifado_id, quantidade)
            )
            importados += 1

        conn.commit()

        # Log da ação
        current_user = get_jwt_identity()
        user = conn.execute(
            'SELECT id FROM usuarios WHERE username = ?', (current_user,)
        ).fetchone()
        if user:
            nome_arquivo = file.filename
            resumo_erros = f' ({len(erros)} erro(s))' if erros else ''
            registrar_log(user['id'], current_user, 'importar', 'estoque',
                          f'Importou {importados} registro(s) de estoque via "{nome_arquivo}"{resumo_erros}')

    msg = f'{importados} produto(s) importado(s) com sucesso.'
    if erros:
        msg += f' {len(erros)} erro(s): ' + '; '.join(erros[:5])
        if len(erros) > 5:
            msg += f' (e mais {len(erros)-5})'

    return response(True, message=msg)

@app.route('/estoque/<int:id>/minimo', methods=['PUT'])
@jwt_required()
def atualizar_estoque_minimo_aba_estoque(id):
    data = request.get_json() or {}
    novo_minimo = data.get('estoque_minimo')
    if novo_minimo is None:
        return response(False, message="Estoque mínimo não fornecido.", status_code=400)
    try:
        with get_connection() as conn:
            # Verifica se o registro de estoque existe
            estoque = conn.execute(
                '''SELECT e.id, e.produto_id, p.nome AS produto_nome, a.nome AS almox_nome
                   FROM estoque e
                   LEFT JOIN produtos p ON e.produto_id = p.id
                   LEFT JOIN almoxarifados a ON e.almoxarifado_id = a.id
                   WHERE e.id = ?''',
                (id,)
            ).fetchone()
            if not estoque:
                return response(False, message="Registro de estoque não encontrado.", status_code=404)

            # Atualiza o estoque_minimo na tabela estoque (por almoxarifado)
            agora = now_iso()
            conn.execute(
                'UPDATE estoque SET estoque_minimo = ?, updated_at = ? WHERE id = ?',
                (novo_minimo, agora, id)
            )
            conn.commit()

            # Gera o LOG DE AUDITORIA
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            if user:
                registrar_log(user['id'], current_user, 'editar', 'estoque',
                              f'Definiu estoque mínimo de "{estoque["produto_nome"]}" em "{estoque["almox_nome"]}" para {novo_minimo}')
            return response(True, message='Estoque mínimo atualizado para este almoxarifado.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

# ============================================================
# MOVIMENTAÇÕES
# ============================================================

@app.route('/movimentacoes', methods=['GET'])
@jwt_required()
def listar_movimentacoes():
    produto_id = request.args.get('produto_id')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    # Paginação
    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    if page < 1:
        page = 1

    try:
        per_page = int(request.args.get('per_page', 50))
    except (TypeError, ValueError):
        per_page = 50
    if per_page < 1:
        per_page = 1
    if per_page > 200:
        per_page = 200

    conditions = []
    params = []

    if produto_id:
        conditions.append('m.produto_id = ?')
        params.append(produto_id)
    if data_inicio:
        conditions.append('m.data >= ?')
        params.append(data_inicio)
    if data_fim:
        conditions.append('m.data <= ?')
        params.append(data_fim)

    where_clause = ''
    if conditions:
        where_clause = ' WHERE ' + ' AND '.join(conditions)

    count_sql = '''
        SELECT COUNT(*) AS count
        FROM movimentacoes m
        LEFT JOIN produtos p ON m.produto_id = p.id
        LEFT JOIN almoxarifados a ON m.almoxarifado_id = a.id
    ''' + where_clause

    data_sql = '''
        SELECT m.id, m.produto_id, m.almoxarifado_id, m.tipo, m.quantidade,
               m.valor_unitario,
               m.quantidade * COALESCE(m.valor_unitario, 0) AS valor_total,
               m.documento, m.tecnico, m.ordem_servico, m.observacao, m.data,
               p.nome AS produto_nome, p.codigo_interno,
               a.nome AS almoxarifado_nome
        FROM movimentacoes m
        LEFT JOIN produtos p ON m.produto_id = p.id
        LEFT JOIN almoxarifados a ON m.almoxarifado_id = a.id
    ''' + where_clause + ' ORDER BY m.data DESC LIMIT ? OFFSET ?'

    offset = (page - 1) * per_page

    try:
        with get_connection() as conn:
            total = conn.execute(count_sql, params).fetchone()['count']
            total_pages = math.ceil(total / per_page) if per_page > 0 else 0

            if page > total_pages and total > 0:
                rows = []
            else:
                rows = conn.execute(data_sql, params + [per_page, offset]).fetchall()

        return response(
            True,
            data=rows_to_dict(rows),
            pagination={
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': total_pages,
            },
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/movimentacoes', methods=['POST'])
@jwt_required()
def criar_movimentacao():
    data = request.get_json(silent=True) or {}

    produto_id = data.get('produto_id')
    almoxarifado_id = data.get('almoxarifado_id')
    tipo = data.get('tipo')
    quantidade = data.get('quantidade')
    valor_unitario = data.get('valor_unitario')
    documento = data.get('documento')
    tecnico = data.get('tecnico')
    ordem_servico = data.get('ordem_servico')
    observacao = data.get('observacao')
    colaborador_id = data.get('colaborador_id')
    lote_corrida = data.get('lote_corrida')
    equipamento_id = data.get('equipamento_id')

    if produto_id is None or tipo is None or quantidade is None:
        return response(False, message='produto_id, tipo e quantidade são obrigatórios.', status_code=400)

    if tipo not in ('entrada', 'saida'):
        return response(False, message="Tipo deve ser 'entrada' ou 'saida'.", status_code=400)

    try:
        with get_connection() as conn:
            # Busca o produto para validar e pegar custo médio atual
            produto = conn.execute(
                'SELECT nome, custo_medio FROM produtos WHERE id = ?', (produto_id,)
            ).fetchone()
            if not produto:
                return response(False, message='Produto não encontrado.', status_code=404)

            custo_medio_atual = produto['custo_medio'] if produto['custo_medio'] is not None else 0

            # Tratamento do valor_unitario conforme o tipo
            if tipo == 'entrada':
                if valor_unitario is not None and valor_unitario > 0:
                    # Recalcula custo médio
                    novo_custo = calcular_custo_medio(conn, produto_id, quantidade, valor_unitario)
                    conn.execute(
                        'UPDATE produtos SET custo_medio = ? WHERE id = ?',
                        (novo_custo, produto_id)
                    )
                else:
                    # Se não informou valor, usa o custo médio atual
                    valor_unitario = custo_medio_atual
            else:  # saida
                if valor_unitario is None or valor_unitario <= 0:
                    valor_unitario = custo_medio_atual

            # Insere a movimentação
            conn.execute('''
                INSERT INTO movimentacoes
                    (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                     valor_unitario, documento, tecnico, ordem_servico, observacao,
                     lote_corrida, equipamento_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                  valor_unitario, documento, tecnico, ordem_servico, observacao,
                  lote_corrida, equipamento_id))

            # Atualiza estoque
            if tipo == 'entrada':
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                       VALUES (?, ?, ?)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + excluded.quantidade''',
                    (produto_id, almoxarifado_id, quantidade)
                )
            else:  # saida
                row = conn.execute(
                    'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                    (produto_id, almoxarifado_id)
                ).fetchone()
                saldo = row['quantidade'] if row else 0
                if saldo < quantidade:
                    return response(False, message='Saldo insuficiente.', status_code=409)

                conn.execute(
                    '''UPDATE estoque
                       SET quantidade = quantidade - ?
                       WHERE produto_id = ? AND almoxarifado_id = ?''',
                    (quantidade, produto_id, almoxarifado_id)
                )

            # Log de auditoria
            current_user = get_jwt_identity()
            user = conn.execute(
                'SELECT id FROM usuarios WHERE username = ?', (current_user,)
            ).fetchone()
            if user:
                registrar_log(
                    user['id'], current_user, tipo, 'movimentacoes',
                    f'{tipo.upper()} {quantidade} un de "{produto["nome"]}" - valor unit: {valor_unitario}',
                    conn=conn
                )

            conn.commit()

            # Busca saldo atualizado
            row = conn.execute(
                'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                (produto_id, almoxarifado_id)
            ).fetchone()
            saldo_atual = row['quantidade'] if row else 0

        return response(
            True,
            data={'saldo_atual': saldo_atual, 'valor_unitario': valor_unitario},
            message='Movimentação registrada com sucesso.'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return response(False, message=str(e), status_code=500)

@app.route('/produtos/<int:id>/codigo-barras', methods=['GET'])
def gerar_codigo_barras(id):
    try:
        with get_connection() as conn:
            produto = conn.execute(
                'SELECT codigo_interno, nome FROM produtos WHERE id = ?',
                (id,)
            ).fetchone()
            if not produto:
                return response(False, message='Produto não encontrado.', status_code=404)

            codigo = produto['codigo_interno']
            if not codigo:
                return response(False, message='Produto sem código interno.', status_code=400)

            # Gera o código de barras no formato CODE128
            CODE128 = barcode.get_barcode_class('code128')
            codigo_barras = CODE128(codigo, writer=ImageWriter())

            # Salva em memória
            buf = io.BytesIO()
            codigo_barras.write(buf)
            buf.seek(0)

            return send_file(buf, mimetype='image/png')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

from markupsafe import escape  # já deve estar no topo, se não estiver, adicione

@app.route('/produtos/etiquetas/lote', methods=['GET'])
def etiquetas_lote():
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401
    
    ids = request.args.get('ids', '')
    if not ids:
        return '<h2>Nenhum produto selecionado</h2>', 400
    
    lista_ids = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not lista_ids:
        return '<h2>IDs inválidos</h2>', 400
    
    placeholders = ','.join(['?'] * len(lista_ids))
    
    with get_connection() as conn:
        produtos = conn.execute(
            f'SELECT id, codigo_interno, nome, unidade FROM produtos WHERE id IN ({placeholders})',
            lista_ids
        ).fetchall()
    
    if not produtos:
        return '<h2>Nenhum produto encontrado</h2>', 404
    
    etiquetas_html = ''
    for p in produtos:
        codigo = p['codigo_interno'] or 'SEMCODIGO'
        nome = p['nome'] or 'Sem nome'
        unidade = p['unidade'] or ''
        etiquetas_html += f'''
        <div class="etiqueta">
            <img src="/produtos/{p['id']}/codigo-barras" alt="Código de Barras">
            <div class="codigo">{escape(codigo)}</div>
            <div class="nome">{escape(nome)}</div>
            <div class="unidade">{escape(unidade)}</div>
        </div>'''
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Etiquetas em Lote - A4 31x17mm</title>
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:Arial, Helvetica, sans-serif; background:#ccc; }}
        .page {{ width:210mm; min-height:297mm; margin:0 auto; background:#fff; padding:8mm 8mm 0 8mm; }}
        .grid {{ display:grid; grid-template-columns:repeat(6, 31mm); grid-template-rows:repeat(16, 17mm); gap:0; justify-content:space-between; }}
        .etiqueta {{ width:31mm; height:17mm; border:0.3mm dashed #999; display:flex; flex-direction:column; align-items:center; justify-content:center; padding:1mm; overflow:hidden; }}
        .etiqueta img {{ max-width:26mm; max-height:7mm; margin-bottom:0.5mm; }}
        .etiqueta .codigo {{ font-size:6px; font-weight:bold; text-align:center; line-height:1.1; }}
        .etiqueta .nome {{ font-size:5.5px; text-align:center; line-height:1.1; color:#333; }}
        .etiqueta .unidade {{ font-size:5px; text-align:center; line-height:1.1; color:#666; }}
        .no-print {{ text-align:center; padding:10px 0; }}
        .no-print button {{ padding:10px 24px; font-size:14px; cursor:pointer; background:#2563eb; color:#fff; border:none; border-radius:6px; margin:0 4px; }}
        @media print {{
            @page {{ size:A4; margin:0; }}
            body {{ background:#fff; }}
            .page {{ width:100%; min-height:auto; padding:8mm 8mm 0 8mm; box-shadow:none; }}
            .no-print {{ display:none; }}
            .etiqueta {{ border:0.2mm solid #000; }}
        }}
    </style>
</head>
<body>
    <div class="no-print">
        <button onclick="window.print()">🖨️ Imprimir ({len(produtos)} etiquetas)</button>
        <button onclick="window.close()">✕ Fechar</button>
        <p style="margin-top:8px;font-size:12px;color:#555;">Formato A4 · 6 colunas × 16 linhas · {len(produtos)} etiqueta(s)</p>
    </div>
    <div class="page">
        <div class="grid">
            {etiquetas_html}
        </div>
    </div>
</body>
</html>'''
    return html

@app.route('/produtos/<int:id>/etiqueta', methods=['GET'])
def etiqueta_produto(id):
    token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return '<h2>Token ausente</h2>', 401
    try:
        decode_token(token)
    except Exception:
        return '<h2>Token inválido ou expirado</h2>', 401
    try:
        with get_connection() as conn:
            produto = conn.execute(
                'SELECT codigo_interno, nome, unidade FROM produtos WHERE id = ?',
                (id,)
            ).fetchone()
            if not produto:
                return '<h2>Produto não encontrado</h2>', 404

            codigo = produto['codigo_interno'] or 'SEMCODIGO'
            nome = produto['nome'] or 'Sem nome'
            unidade = produto['unidade'] or ''

            html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Etiqueta - {nome}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: Arial, sans-serif;
            display: flex; flex-direction: column; align-items: center;
            justify-content: center; min-height: 100vh; padding: 20px;
        }}
        .etiqueta {{
            width: 280px; padding: 16px; border: 2px solid #333;
            border-radius: 8px; text-align: center; background: #fff;
        }}
        .etiqueta img {{ width: 220px; height: auto; margin-bottom: 8px; }}
        .etiqueta .nome {{ font-size: 14px; font-weight: bold; margin-bottom: 4px; }}
        .etiqueta .info {{ font-size: 12px; color: #555; }}
        .btn-imprimir {{
            margin-top: 20px; padding: 12px 24px; font-size: 16px;
            cursor: pointer; background: #2563eb; color: #fff;
            border: none; border-radius: 6px;
        }}
        @media print {{
            .btn-imprimir {{ display: none; }}
            body {{ padding: 0; }}
            .etiqueta {{ border: 1px solid #000; }}
        }}
    </style>
</head>
<body>
    <button class="btn-imprimir" onclick="window.print()">🖨️ Imprimir Etiqueta</button>
    <div class="etiqueta">
        <img src="/produtos/{id}/codigo-barras" alt="Código de Barras">
        <div class="nome">{nome}</div>
        <div class="info">Cód: {codigo} | {unidade}</div>
    </div>
</body>
</html>'''
            return html
    except Exception as e:
        return f'<h2>Erro: {str(e)}</h2>', 500

# ============================================================
# MOVIMENTAÇÃO RÁPIDA (SCANNER)
# ============================================================
@app.route('/colaboradores/busca', methods=['GET'])
@jwt_required()
def buscar_colaborador_por_codigo():
    codigo = request.args.get('codigo_barras', '').strip()
    if not codigo:
        return response(False, message='Informe o código de barras ou matrícula.', status_code=400)
    try:
        with get_connection() as conn:
            # Primeiro tenta por código de barras
            row = conn.execute(
                "SELECT id, nome, matricula, setor FROM colaboradores WHERE codigo_barras = ? AND status = 'ativo'",
                (codigo,)
            ).fetchone()
            
            # Se não achou, tenta por matrícula
            if not row:
                row = conn.execute(
                    "SELECT id, nome, matricula, setor FROM colaboradores WHERE matricula = ? AND status = 'ativo'",
                    (codigo,)
                ).fetchone()

            if not row:
                return response(False, message='Colaborador não encontrado ou inativo. Verifique o código ou matrícula.', status_code=404)
            
            return response(True, data=dict(row))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/produtos/busca', methods=['GET'])
@jwt_required()
def buscar_produto_por_codigo():
    codigo = request.args.get('codigo', '').strip()
    if not codigo:
        return response(False, message='Informe o código do produto.', status_code=400)
    try:
        with get_connection() as conn:
            row = conn.execute('''
                SELECT p.id, p.nome, p.codigo_interno, p.categoria, p.natureza,
                       p.custo_medio,
                       COALESCE(SUM(e.quantidade), 0) AS quantidade_estoque
                FROM produtos p
                LEFT JOIN estoque e ON e.produto_id = p.id
                WHERE p.codigo_interno = ? AND p.ativo = 1
                GROUP BY p.id
                LIMIT 1
            ''', (codigo,)).fetchone()
            if not row:
                return response(False, message='Produto não encontrado ou inativo.', status_code=404)
            return response(True, data=dict(row))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/movimentacoes/retirada-rapida', methods=['POST'])
@jwt_required()
def retirada_rapida():
    data = request.get_json(silent=True) or {}
    colaborador_id = data.get('colaborador_id')
    almoxarifado_id = data.get('almoxarifado_id')
    itens = data.get('itens', [])
    if not colaborador_id or not itens:
        return response(False, message='colaborador_id e itens são obrigatórios.', status_code=400)
    if not almoxarifado_id:
        # Pega o primeiro almoxarifado como padrão
        with get_connection() as conn:
            row = conn.execute('SELECT id FROM almoxarifados ORDER BY id LIMIT 1').fetchone()
            if row:
                almoxarifado_id = row['id']
            else:
                return response(False, message='Nenhum almoxarifado cadastrado.', status_code=400)
    current_user = get_jwt_identity()
    try:
        with get_connection() as conn:
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            colaborador = conn.execute('SELECT nome FROM colaboradores WHERE id = ?', (colaborador_id,)).fetchone()
            if not colaborador:
                return response(False, message='Colaborador não encontrado.', status_code=404)

            resultados = []
            itens_inseridos = 0

            for item in itens:
                produto_id = item.get('produto_id')
                quantidade = item.get('quantidade', 1)
                if not produto_id:
                    continue

                produto = conn.execute(
                    'SELECT nome, natureza, custo_medio FROM produtos WHERE id = ?',
                    (produto_id,)
                ).fetchone()
                if not produto:
                    continue

                natureza = produto['natureza'] or 'consumivel'
                valor_unitario = produto['custo_medio'] or 0

                # Verifica saldo
                saldo_row = conn.execute(
                    'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                    (produto_id, almoxarifado_id)
                ).fetchone()
                saldo = saldo_row['quantidade'] if saldo_row else 0
                if saldo < quantidade:
                    return response(
                        False,
                        message=f'Saldo insuficiente de "{produto["nome"]}": disponível {saldo}, solicitado {quantidade}.',
                        status_code=409
                    )

                # Insere movimentação
                conn.execute('''
                    INSERT INTO movimentacoes
                        (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                         valor_unitario, natureza, devolvido)
                    VALUES (?, ?, ?, 'saida', ?, ?, ?, ?)
                ''', (produto_id, almoxarifado_id, colaborador_id, quantidade,
                      valor_unitario, natureza, False))

                # Atualiza estoque
                conn.execute(
                    'UPDATE estoque SET quantidade = quantidade - ? WHERE produto_id = ? AND almoxarifado_id = ?',
                    (quantidade, produto_id, almoxarifado_id)
                )

                resultados.append({
                    'produto_id': produto_id,
                    'produto_nome': produto['nome'],
                    'natureza': natureza,
                    'quantidade': quantidade
                })
                itens_inseridos += 1

            # Log
            if user and itens_inseridos > 0:
                registrar_log(
                    user['id'], current_user, 'saida', 'retirada_rapida',
                    f'Retirada rápida de {itens_inseridos} item(ns) por "{colaborador["nome"]}"',
                    conn=conn
                )

            conn.commit()

        return response(
            True,
            data={'itens': resultados, 'total_itens': itens_inseridos},
            message=f'{itens_inseridos} item(ns) retirado(s) com sucesso.'
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/movimentacoes/devolucao-unidade', methods=['POST'])
@jwt_required()
def devolucao_unidade():
    data = request.get_json(silent=True) or {}
    colaborador_id = data.get('colaborador_id')
    produto_id = data.get('produto_id')
    quantidade = data.get('quantidade', 1)
    almoxarifado_id = data.get('almoxarifado_id')

    if not colaborador_id or not produto_id:
        return response(False, message='colaborador_id e produto_id são obrigatórios.', status_code=400)

    try:
        with get_connection() as conn:
            # Localiza movimentação em aberto (não devolvida)
            mov = conn.execute('''
                SELECT m.id, m.natureza, m.quantidade
                FROM movimentacoes m
                WHERE m.colaborador_id = ?
                  AND m.produto_id = ?
                  AND m.tipo = 'saida'
                  AND m.natureza = 'emprestimo'
                  AND m.devolvido = FALSE
                ORDER BY m.data DESC
                LIMIT 1
            ''', (colaborador_id, produto_id)).fetchone()

            if not mov:
                return response(
                    False,
                    message='Nenhum empréstimo pendente encontrado para este colaborador/produto.',
                    status_code=404
                )

            if not almoxarifado_id:
                row = conn.execute('SELECT id FROM almoxarifados ORDER BY id LIMIT 1').fetchone()
                almoxarifado_id = row['id'] if row else None

            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            produto = conn.execute('SELECT nome FROM produtos WHERE id = ?', (produto_id,)).fetchone()

            # Marca como devolvido
            conn.execute('''
                UPDATE movimentacoes
                SET devolvido = TRUE, data_devolucao = ?
                WHERE id = ?
            ''', (now_iso(), mov['id']))

            # Registra entrada de devolução no estoque
            if almoxarifado_id:
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                       VALUES (?, ?, ?)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + excluded.quantidade''',
                    (produto_id, almoxarifado_id, quantidade)
                )

                # Também registra como movimentação de entrada (para o histórico)
                conn.execute('''
                    INSERT INTO movimentacoes
                        (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                         valor_unitario, natureza, devolvido, observacao)
                    VALUES (?, ?, ?, 'entrada', ?, ?, 'emprestimo', TRUE, ?)
                ''', (produto_id, almoxarifado_id, colaborador_id, quantidade,
                      0, f'Devolução do empréstimo #{mov["id"]}'))

            # Log
            if user and produto:
                registrar_log(
                    user['id'], current_user, 'entrada', 'devolucao_rapida',
                    f'Devolução de {quantidade} un de "{produto["nome"]}" pelo colaborador',
                    conn=conn
                )

            conn.commit()

        return response(
            True,
            data={'movimentacao_original_id': mov['id']},
            message='Devolução registrada com sucesso.'
        )
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/movimentacoes/pendencias', methods=['GET'])
@jwt_required()
def listar_pendencias():
    colaborador_id = request.args.get('colaborador_id')
    if not colaborador_id:
        return response(False, message='colaborador_id é obrigatório.', status_code=400)
    try:
        with get_connection() as conn:
            rows = conn.execute('''
                SELECT m.id, m.produto_id, m.quantidade, m.data as data_retirada,
                       p.nome AS produto_nome, p.codigo_interno
                FROM movimentacoes m
                LEFT JOIN produtos p ON m.produto_id = p.id
                WHERE m.colaborador_id = ?
                  AND m.tipo = 'saida'
                  AND m.natureza = 'emprestimo'
                  AND m.devolvido = FALSE
                ORDER BY m.data DESC
            ''', (colaborador_id,)).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

# ============================================================
# EMPRÉSTIMOS
# ============================================================
@app.route('/emprestimos', methods=['GET'])
@jwt_required()
def listar_emprestimos():
    try:
        with get_connection() as conn:
            rows = conn.execute('''
                SELECT ep.*, u.numero_serie, p.nome as produto_nome,
                       c.nome as colaborador_nome
                FROM emprestimos ep
                LEFT JOIN unidades u ON ep.unidade_id = u.id
                LEFT JOIN produtos p ON u.produto_id = p.id
                LEFT JOIN colaboradores c ON ep.colaborador_id = c.id
                ORDER BY ep.data_emprestimo DESC
            ''').fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/emprestimos', methods=['POST'])
@jwt_required()
def criar_emprestimo():
    data = request.get_json() or {}
    unidade_id = data.get('unidade_id')
    colaborador_id = data.get('colaborador_id')
    data_emprestimo = data.get('data_emprestimo')
    observacao = data.get('observacao', '')
    tipo = data.get('tipo', 'emprestimo')
    if not unidade_id or not colaborador_id:
        return response(False, message='Unidade e colaborador são obrigatórios.', status_code=400)
    try:
        with get_connection() as conn:
            conn.execute(
                '''INSERT INTO emprestimos
                   (unidade_id, colaborador_id, data_emprestimo, observacao, status, tipo, created_at)
                   VALUES (?, ?, ?, ?, 'ativo', ?, ?)''',
                (unidade_id, colaborador_id, data_emprestimo, observacao, tipo, now_iso())
            )
            conn.execute(
                'UPDATE unidades SET status = ? WHERE id = ?',
                ('emprestado', unidade_id)
            )
            conn.commit()
        return response(True, message='Empréstimo registrado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/emprestimos/devolver', methods=['POST'])
@jwt_required()
def devolver_emprestimo():
    data = request.get_json() or {}
    unidade_id = data.get('unidade_id')
    observacao = data.get('observacao', '')
    if not unidade_id:
        return response(False, message='Unidade é obrigatória.', status_code=400)
    try:
        with get_connection() as conn:
            emp = conn.execute(
                'SELECT id FROM emprestimos WHERE unidade_id = ? AND status = ? ORDER BY id DESC LIMIT 1',
                (unidade_id, 'ativo')
            ).fetchone()
            if emp:
                conn.execute(
                    'UPDATE emprestimos SET status = ?, data_devolucao = ? WHERE id = ?',
                    ('devolvido', now_iso(), emp['id'])
                )
            conn.execute(
                'UPDATE unidades SET status = ? WHERE id = ?',
                ('disponivel', unidade_id)
            )
            conn.commit()
        return response(True, message='Devolução registrada com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/emprestimos/manutencao', methods=['POST'])
@jwt_required()
def manutencao_emprestimo():
    data = request.get_json() or {}
    unidade_id = data.get('unidade_id')
    observacao = data.get('observacao', '')
    if not unidade_id:
        return response(False, message='Unidade é obrigatória.', status_code=400)
    try:
        with get_connection() as conn:
            emp = conn.execute(
                'SELECT id FROM emprestimos WHERE unidade_id = ? AND status = ? ORDER BY id DESC LIMIT 1',
                (unidade_id, 'ativo')
            ).fetchone()
            if emp:
                conn.execute(
                    'UPDATE emprestimos SET status = ?, data_devolucao = ? WHERE id = ?',
                    ('manutencao', now_iso(), emp['id'])
                )
            conn.execute(
                'UPDATE unidades SET status = ? WHERE id = ?',
                ('manutencao', unidade_id)
            )
            conn.commit()
        return response(True, message='Unidade enviada para manutenção.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

# ============================================================
# RELATÓRIOS
# ============================================================

@app.route('/relatorios/estoque', methods=['GET'])
@jwt_required()
def relatorio_estoque():
    try:
        tipo = request.args.get('tipo', 'baixo')

        base_sql = '''
            SELECT p.id, p.nome, p.codigo_fabricante, p.descricao,
                   COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0) as estoque_minimo,
                   COALESCE(e.saldo, 0) as saldo,
                   (SELECT pc.status FROM pedidos_compra pc WHERE pc.produto_id = p.id AND pc.status IN ('aberto','em_compra') ORDER BY pc.id DESC LIMIT 1) as pedido_status,
                   (SELECT pc.data_prevista_chegada FROM pedidos_compra pc WHERE pc.produto_id = p.id AND pc.status IN ('aberto','em_compra') ORDER BY pc.id DESC LIMIT 1) as data_prevista,
                   (SELECT pc.id FROM pedidos_compra pc WHERE pc.produto_id = p.id AND pc.status IN ('aberto','em_compra') ORDER BY pc.id DESC LIMIT 1) as pedido_id
            FROM produtos p
            LEFT JOIN (
                SELECT produto_id, SUM(quantidade) as saldo,
                       SUM(COALESCE(estoque_minimo, 0)) as total_estoque_minimo
                FROM estoque
                GROUP BY produto_id
            ) e ON p.id = e.produto_id
        '''

        if tipo == 'zero':
            sql = base_sql + ' WHERE COALESCE(e.saldo, 0) = 0 ORDER BY p.nome'
        elif tipo == 'critico':
            sql = base_sql + ' WHERE COALESCE(e.saldo, 0) > 0 AND COALESCE(e.saldo, 0) <= COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0) / 2 ORDER BY (COALESCE(e.saldo, 0) - COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0)) ASC'
        elif tipo == 'todos':
            sql = base_sql + ' WHERE COALESCE(e.saldo, 0) <= COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0) ORDER BY (COALESCE(e.saldo, 0) - COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0)) ASC'
        else:
            sql = base_sql + ' WHERE COALESCE(e.saldo, 0) > 0 AND COALESCE(e.saldo, 0) <= COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0) AND COALESCE(e.saldo, 0) > COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0) / 2 ORDER BY (COALESCE(e.saldo, 0) - COALESCE(e.total_estoque_minimo, p.estoque_minimo, 0)) ASC'

        with get_connection() as conn:
            rows = conn.execute(sql).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


@app.route('/relatorios/movimentacoes-periodo', methods=['GET'])
@jwt_required()
def relatorio_movimentacoes_periodo():
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    sql = '''SELECT m.*, p.nome as produto_nome, p.codigo_interno, p.categoria,
                    c.nome as colaborador_nome, a.nome as almoxarifado_nome
             FROM movimentacoes m
             LEFT JOIN produtos p ON m.produto_id = p.id
             LEFT JOIN colaboradores c ON m.colaborador_id = c.id
             LEFT JOIN almoxarifados a ON m.almoxarifado_id = a.id
             WHERE 1=1'''
    params = []

    if data_inicio:
        sql += ' AND m.data >= ?'
        params.append(data_inicio)
    if data_fim:
        sql += ' AND m.data <= ?'
        params.append(data_fim)

    sql += ' ORDER BY m.data DESC'

    try:
        with get_connection() as conn:
            rows = conn.execute(sql, params).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)


# ============================================================
# DASHBOARD
# ============================================================

@app.route('/dashboard', methods=['GET'])
@jwt_required()
def dashboard():
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        with get_connection() as conn:
                        # Gastos com consumíveis (saídas de produtos que NÃO controlam depreciação)
            query_gastos = '''SELECT COALESCE(SUM(m.quantidade * COALESCE(p.custo_medio, 0)), 0) as total
                   FROM movimentacoes m
                   JOIN produtos p ON m.produto_id = p.id
                   WHERE m.tipo = 'saida' AND COALESCE(p.controla_depreciacao, 0) = 0'''
            params_gastos = []
            if data_inicio:
                query_gastos += ' AND m.data >= ?'
                params_gastos.append(data_inicio)
            if data_fim:
                query_gastos += ' AND m.data <= ?'
                params_gastos.append(data_fim)
            gastos_consumiveis = conn.execute(query_gastos, params_gastos).fetchone()['total']

            # Aquisições de ativos (entradas de produtos QUE controlam depreciação)
            query_aquis = '''SELECT COALESCE(SUM(m.quantidade * COALESCE(m.valor_unitario, 0)), 0) as total
                   FROM movimentacoes m
                   JOIN produtos p ON m.produto_id = p.id
                   WHERE m.tipo = 'entrada' AND COALESCE(p.controla_depreciacao, 0) = 1'''
            params_aquis = []
            if data_inicio:
                query_aquis += ' AND m.data >= ?'
                params_aquis.append(data_inicio)
            if data_fim:
                query_aquis += ' AND m.data <= ?'
                params_aquis.append(data_fim)
            aquisicoes_ativos = conn.execute(query_aquis, params_aquis).fetchone()['total']

            estoque_baixo = conn.execute(
                '''SELECT COUNT(*) as total FROM estoque e
                   WHERE e.quantidade < COALESCE(e.estoque_minimo, 
                       (SELECT p.estoque_minimo FROM produtos p WHERE p.id = e.produto_id), 0)'''
            ).fetchone()['total']
            equip_manutencao = conn.execute(
                "SELECT COUNT(*) as total FROM unidades WHERE status_manutencao = 'em_manutencao'"
            ).fetchone()['total']
            calibracoes_vencer = conn.execute(
                '''SELECT COUNT(*) as total FROM unidades
                   WHERE requer_calibracao = 1
                     AND data_validade_calibracao IS NOT NULL
                     AND julianday(data_validade_calibracao) - julianday('now') BETWEEN 0 AND 30'''
            ).fetchone()['total']
            maiores_consumidores = conn.execute(
                '''SELECT c.nome,
                          COALESCE(SUM(m.quantidade), 0) as quantidade_total,
                          COALESCE(SUM(m.quantidade * COALESCE(m.valor_unitario, 0)), 0) as valor_total
                   FROM movimentacoes m
                   LEFT JOIN colaboradores c ON m.colaborador_id = c.id
                   WHERE m.tipo = 'saida'
                   GROUP BY c.nome, m.colaborador_id
                   ORDER BY quantidade_total DESC
                   LIMIT 10'''
            ).fetchall()
            gastos_por_tipo = conn.execute(
                '''SELECT p.categoria,
                          COALESCE(SUM(m.quantidade * COALESCE(m.valor_unitario, 0)), 0) as total
                   FROM movimentacoes m
                   JOIN produtos p ON m.produto_id = p.id
                   WHERE m.tipo = 'entrada' AND m.valor_unitario > 0
                   GROUP BY p.categoria'''
            ).fetchall()
            equipamentos_manutencao_lista = conn.execute(
                '''SELECT u.id, u.tag, u.numero_serie, p.nome as produto_nome
                   FROM unidades u
                   LEFT JOIN produtos p ON u.produto_id = p.id
                   WHERE u.status_manutencao = 'em_manutencao' '''
            ).fetchall()
            calibracoes_lista = conn.execute(
                '''SELECT u.id, u.tag, u.numero_serie, p.nome as produto_nome,
                          u.data_validade_calibracao,
                          CAST(julianday(u.data_validade_calibracao) - julianday('now') AS INTEGER) as dias_restantes
                   FROM unidades u
                   LEFT JOIN produtos p ON u.produto_id = p.id
                   WHERE u.requer_calibracao = 1
                     AND u.data_validade_calibracao IS NOT NULL
                     AND julianday(u.data_validade_calibracao) - julianday('now') BETWEEN 0 AND 30
                   ORDER BY u.data_validade_calibracao'''
            ).fetchall()
            # Valor do estoque de consumíveis (produtos que não controlam depreciação)
            valor_estoque_consumiveis = conn.execute(
                '''SELECT COALESCE(SUM(e.quantidade * COALESCE(p.custo_medio, 0)), 0) as total
                   FROM estoque e
                   JOIN produtos p ON e.produto_id = p.id
                   WHERE COALESCE(p.controla_depreciacao, 0) = 0'''
            ).fetchone()['total']
            # Total de unidades ativas para depreciação
            total_unidades_ativas = conn.execute(
                "SELECT COUNT(*) as total FROM unidades WHERE status_depreciacao = 'ativo'"
            ).fetchone()['total']
            # Cálculo dos totais de depreciação (VLC) via função do módulo database
            depreciacao_totais = calcular_vlc_total(conn)
                        # Total de gastos com manutenção no período
            query_manutencao = '''SELECT COALESCE(SUM(COALESCE(custo, 0)), 0) as total
                   FROM manutencoes_unidades
                   WHERE 1=1'''
            params_manutencao = []
            if data_inicio:
                query_manutencao += ' AND data_envio >= ?'
                params_manutencao.append(data_inicio)
            if data_fim:
                query_manutencao += ' AND data_envio <= ?'
                params_manutencao.append(data_fim)
            total_gastos_manutencao = conn.execute(query_manutencao, params_manutencao).fetchone()['total']
        return response(True, data={
            'gastos_consumiveis': gastos_consumiveis,
            'aquisicoes_ativos': aquisicoes_ativos,
            'estoque_baixo': estoque_baixo,
            'equip_manutencao': equip_manutencao,
            'calibracoes_vencer': calibracoes_vencer,
            'maiores_consumidores': rows_to_dict(maiores_consumidores),
            'gastos_por_tipo': rows_to_dict(gastos_por_tipo),
            'equipamentos_manutencao_lista': rows_to_dict(equipamentos_manutencao_lista),
            'calibracoes_lista': rows_to_dict(calibracoes_lista),
            'valor_estoque_consumiveis': valor_estoque_consumiveis,
            'total_unidades_ativas': total_unidades_ativas,
            'valor_aquisicao_total': depreciacao_totais.get('valor_aquisicao_total', 0),
            'vlc_total': depreciacao_totais.get('vlc_total', 0),
            'valor_residual_total': depreciacao_totais.get('valor_residual_total', 0),
            'depreciacao_acumulada_total': depreciacao_totais.get('depreciacao_acumulada_total', 0),
            'total_gastos_manutencao': total_gastos_manutencao
        })
    except Exception as e:
        return response(False, message=str(e), status_code=500)


# ============================================================
# FUNÇÃO PARA CRIAR USUÁRIO VIA TERMINAL
# ============================================================

def criar_usuario(username, password):
    if not username or not password:
        print('Username e password são obrigatórios.')
        return

    try:
        with get_connection() as conn:
            row = conn.execute('SELECT id FROM usuarios WHERE username = ?', (username,)).fetchone()
            if row:
                print(f'Usuário "{username}" já existe.')
                return

            conn.execute(
                'INSERT INTO usuarios (username, password, created_at) VALUES (?, ?, ?)',
                (username, generate_password_hash(password), now_iso())
            )
            conn.commit()
        print(f'Usuário "{username}" criado com sucesso.')
    except Exception as e:
        print(f'Erro ao criar usuário: {e}')

# ============================================================
# PEDIDOS DE COMPRA (multi-item)
# ============================================================
@app.route('/pedidos', methods=['GET'])
@jwt_required()
def listar_pedidos():
    try:
        status = request.args.get('status')
        base_sql = '''
            SELECT p.*,
                   COUNT(pi.id) as total_itens,
                   SUM(pi.quantidade_solicitada - COALESCE(pi.quantidade_transferida, 0)) as qtd_total,
                   SUM(COALESCE(pi.quantidade_transferida, 0)) as qtd_transferida,
                   COALESCE(SUM(pi.preco_unitario * (pi.quantidade_solicitada - COALESCE(pi.quantidade_transferida, 0))), 0) as valor_total
            FROM pedidos p
            LEFT JOIN pedidos_itens pi ON p.id = pi.pedido_id
        '''
        if status:
            sql = base_sql + " WHERE p.status = ? GROUP BY p.id ORDER BY CASE p.status WHEN 'aberto' THEN 1 WHEN 'em_compra' THEN 2 WHEN 'comprado' THEN 3 WHEN 'recebido' THEN 4 ELSE 5 END, p.created_at DESC"
        else:
            sql = base_sql + " GROUP BY p.id ORDER BY CASE p.status WHEN 'aberto' THEN 1 WHEN 'em_compra' THEN 2 WHEN 'comprado' THEN 3 WHEN 'recebido' THEN 4 ELSE 5 END, p.created_at DESC"
        with get_connection() as conn:
            if status:
                rows = conn.execute(sql, (status,)).fetchall()
            else:
                rows = conn.execute(sql).fetchall()
        return response(True, data=rows_to_dict(rows))
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/pedidos/<int:id>', methods=['GET'])
@jwt_required()
def obter_pedido(id):
    try:
        with get_connection() as conn:
            pedido = conn.execute('''
                SELECT p.*,
                       COUNT(pi.id) as total_itens,
                       SUM(pi.quantidade_solicitada) as qtd_total,
                       COALESCE(SUM(pi.preco_unitario * pi.quantidade_solicitada), 0) as valor_total
                FROM pedidos p
                LEFT JOIN pedidos_itens pi ON p.id = pi.pedido_id
                WHERE p.id = ?
                GROUP BY p.id
            ''', (id,)).fetchone()
            if not pedido:
                return response(False, message='Pedido não encontrado.', status_code=404)
            itens = conn.execute('''
                SELECT pi.*, pr.nome as produto_nome, pr.codigo_fabricante, pr.codigo_interno,
                       (pi.quantidade_solicitada - COALESCE(pi.quantidade_transferida, 0)) as saldo
                FROM pedidos_itens pi
                LEFT JOIN produtos pr ON pi.produto_id = pr.id
                WHERE pi.pedido_id = ?
            ''', (id,)).fetchall()
            result = dict(pedido)
            result['itens'] = rows_to_dict(itens)
        return response(True, data=result)
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/pedidos', methods=['POST'])
@jwt_required()
def criar_pedido():
    try:
        data = request.get_json(silent=True) or {}
        fornecedor = data.get('fornecedor')
        solicitante = data.get('solicitante')
        observacao = data.get('observacao')
        itens = data.get('itens', [])
        if not itens:
            return response(False, message='É necessário informar pelo menos 1 item.', status_code=400)
        current_user = get_jwt_identity()
        with get_connection() as conn:
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            pedido_id = conn.execute('''
                INSERT INTO pedidos
                    (fornecedor, solicitante, observacao, status, data_abertura, data_pedido, created_at, updated_at)
                VALUES (?, ?, ?, 'aberto', ?, ?, ?, ?)
                RETURNING id
            ''', (fornecedor, solicitante, observacao, now_iso(), now_iso(), now_iso(), now_iso())).fetchone()['id']
            for item in itens:
                conn.execute('''
                    INSERT INTO pedidos_itens (pedido_id, produto_id, quantidade_solicitada, preco_unitario)
                    VALUES (?, ?, ?, ?)
                ''', (pedido_id, item['produto_id'], item['quantidade_solicitada'], item.get('preco_unitario')))
            registrar_log(
                user['id'], current_user, 'criar', 'pedidos',
                f'Criou pedido #{pedido_id} com {len(itens)} item(ns)',
                conn=conn
            )
            conn.commit()
        return response(True, data={'id': pedido_id}, message='Pedido criado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/pedidos/<int:id>', methods=['PUT'])
@jwt_required()
def atualizar_pedido(id):
    try:
        data = request.get_json(silent=True) or {}
        with get_connection() as conn:
            existente = conn.execute('SELECT id FROM pedidos WHERE id = ?', (id,)).fetchone()
            if not existente:
                return response(False, message='Pedido não encontrado.', status_code=404)
            campos_pedido = ['fornecedor', 'solicitante', 'observacao']
            valores = {c: data[c] for c in campos_pedido if c in data}
            if data.get('status'):
                valores['status'] = data['status']
                # Atualiza timestamp do status
                if data['status'] == 'em_compra':
                    valores['data_em_compra'] = now_iso()
                elif data['status'] == 'comprado':
                    valores['data_comprado'] = now_iso()
                elif data['status'] == 'recebido':
                    valores['data_recebido'] = now_iso()
            if valores:
                valores['updated_at'] = now_iso()
                set_clause = ', '.join([f'{c} = ?' for c in valores.keys()])
                params = tuple(valores.values()) + (id,)
                conn.execute(f'UPDATE pedidos SET {set_clause} WHERE id = ?', params)
            # Se veio lista de itens, substitui
            if data.get('itens') is not None:
                conn.execute('DELETE FROM pedidos_itens WHERE pedido_id = ?', (id,))
                for item in data['itens']:
                    conn.execute('''
                        INSERT INTO pedidos_itens (pedido_id, produto_id, quantidade_solicitada, preco_unitario)
                        VALUES (?, ?, ?, ?)
                    ''', (id, item['produto_id'], item['quantidade_solicitada'], item.get('preco_unitario')))
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            registrar_log(
                user['id'], current_user, 'atualizar', 'pedidos',
                f'Atualizou pedido #{id}',
                conn=conn
            )
            conn.commit()
        return response(True, message='Pedido atualizado com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/pedidos/<int:id>', methods=['DELETE'])
@jwt_required()
@perfil_required('admin', 'operador')
def deletar_pedido(id):
    try:
        with get_connection() as conn:
            cur = conn.execute('DELETE FROM pedidos WHERE id = ?', (id,))
            if cur.rowcount == 0:
                return response(False, message='Pedido não encontrado.', status_code=404)
            current_user = get_jwt_identity()
            user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            registrar_log(
                user['id'], current_user, 'deletar', 'pedidos',
                f'Deletou pedido #{id}',
                conn=conn
            )
            conn.commit()
        return response(True, message='Pedido excluído com sucesso.')
    except Exception as e:
        return response(False, message=str(e), status_code=500)

@app.route('/pedidos/<int:id>/transferir', methods=['POST'])
@jwt_required()
@perfil_required('admin', 'operador')
def transferir_itens_pedido(id):
    try:
        data = request.get_json(silent=True) or {}
        itens = data.get('itens', [])

        if not itens:
            return response(False, message='Nenhum item informado para transferência.', status_code=400)

        current_user = get_jwt_identity()

        with get_connection() as conn:
            # Validar pedido
            pedido = conn.execute('SELECT id, status FROM pedidos WHERE id = ?', (id,)).fetchone()
            if not pedido:
                return response(False, message='Pedido não encontrado.', status_code=404)
            if pedido['status'] != 'aberto':
                return response(False, message='Apenas pedidos em aberto podem ter itens transferidos.', status_code=400)

            # Buscar itens do pedido
            itens_pedido = {
                row['produto_id']: dict(row)
                for row in conn.execute(
                    '''SELECT pi.*, pr.nome as produto_nome
                       FROM pedidos_itens pi
                       JOIN produtos pr ON pi.produto_id = pr.id
                       WHERE pi.pedido_id = ?''', (id,)
                ).fetchall()
            }

            transferencias_criadas = []

            for item in itens:
                produto_id = item.get('produto_id')
                quantidade = item.get('quantidade', 0)
                origem_id = item.get('almoxarifado_origem_id')
                destino_id = item.get('almoxarifado_destino_id')

                if not all([produto_id, quantidade, origem_id, destino_id]):
                    return response(False, message='Cada item deve ter: produto_id, quantidade, almoxarifado_origem_id, almoxarifado_destino_id.', status_code=400)

                try:
                    quantidade = int(quantidade)
                except (TypeError, ValueError):
                    return response(False, message=f'Quantidade inválida para produto {produto_id}.', status_code=400)

                if quantidade <= 0:
                    return response(False, message=f'Quantidade deve ser maior que zero.', status_code=400)

                if origem_id == destino_id:
                    return response(False, message=f'Origem e destino devem ser diferentes.', status_code=400)

                # Validar que o item pertence ao pedido
                if produto_id not in itens_pedido:
                    return response(False, message=f'Produto {produto_id} não pertence a este pedido.', status_code=400)

                pi = itens_pedido[produto_id]
                qtd_transferida = pi.get('quantidade_transferida', 0) or 0
                saldo_disponivel = pi['quantidade_solicitada'] - qtd_transferida

                if quantidade > saldo_disponivel:
                    return response(False,
                        message=f'{pi["produto_nome"]}: transferência excede saldo disponível ({saldo_disponivel}).',
                        status_code=400)

                # Validar almoxarifados
                origem = conn.execute('SELECT id, nome FROM almoxarifados WHERE id = ?', (origem_id,)).fetchone()
                if not origem:
                    return response(False, message=f'Almoxarifado de origem não encontrado.', status_code=404)
                destino = conn.execute('SELECT id, nome FROM almoxarifados WHERE id = ?', (destino_id,)).fetchone()
                if not destino:
                    return response(False, message=f'Almoxarifado de destino não encontrado.', status_code=404)

                # Verificar saldo no estoque
                saldo_row = conn.execute(
                    'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                    (produto_id, origem_id)
                ).fetchone()
                saldo_atual = saldo_row['quantidade'] if saldo_row else 0

                if saldo_atual < quantidade:
                    return response(False,
                        message=f'Saldo insuficiente de {pi["produto_nome"]} em {origem["nome"]}.',
                        status_code=409)

                # 1. Dar baixa no estoque de origem
                conn.execute(
                    'UPDATE estoque SET quantidade = quantidade - ? WHERE produto_id = ? AND almoxarifado_id = ?',
                    (quantidade, produto_id, origem_id)
                )

                # 2. Criar transferencia
                conn.execute(
                    '''INSERT INTO transferencias
                       (produto_id, almoxarifado_origem_id, almoxarifado_destino_id,
                        quantidade_total, quantidade_recebida, valor_unitario, status,
                        documento, tecnico, ordem_servico, observacao, data_envio, enviado_por)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (produto_id, origem_id, destino_id,
                     quantidade, 0, None, 'enviada',
                     f'Pedido #{id}', None, None,
                     f'Transferência automática do Pedido #{id}', now_iso(), current_user)
                )

                trow = conn.execute('SELECT id FROM transferencias ORDER BY id DESC LIMIT 1').fetchone()
                transfer_id = trow['id']

                # 3. Registrar movimentação de saída
                conn.execute(
                    '''INSERT INTO movimentacoes
                       (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                        valor_unitario, documento, tecnico, ordem_servico, observacao, data)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (produto_id, origem_id, None, 'saida', quantidade,
                     None, f'Pedido #{id}', None, None,
                     f'Transferência #{transfer_id} para {destino["nome"]} (Pedido #{id})', now_iso())
                )

                # 4. Atualizar quantidade_transferida no pedido
                conn.execute(
                    'UPDATE pedidos_itens SET quantidade_transferida = COALESCE(quantidade_transferida, 0) + ? WHERE id = ?',
                    (quantidade, pi['id'])
                )

                transferencias_criadas.append({
                    'transfer_id': transfer_id,
                    'produto_id': produto_id,
                    'produto_nome': pi['produto_nome'],
                    'quantidade': quantidade,
                    'origem': origem['nome'],
                    'destino': destino['nome']
                })

            # Registrar log
            user_row = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            user_id = user_row['id'] if user_row else None
            registrar_log(user_id, current_user, 'transferir', 'pedidos',
                f'Transferência de {len(transferencias_criadas)} itens do Pedido #{id}',
                conn=conn)

            # Verificar se todos os itens do pedido foram totalmente transferidos
            restante = conn.execute('''
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN COALESCE(quantidade_transferida, 0) < quantidade_solicitada THEN 1 ELSE 0 END) as pendentes
                FROM pedidos_itens WHERE pedido_id = ?
            ''', (id,)).fetchone()
            if restante and restante['pendentes'] == 0:
                conn.execute('UPDATE pedidos SET status = ?, data_recebido = ? WHERE id = ?',
                             ('recebido', now_iso(), id))

            conn.commit()

        return response(True, data={
            'transferencias': transferencias_criadas,
            'total_itens': len(transferencias_criadas)
        }, message=f'{len(transferencias_criadas)} itens transferidos com sucesso.')

    except Exception:
        import traceback; traceback.print_exc()
        return response(False, message='Erro interno ao transferir itens do pedido.', status_code=500)

@app.route('/pedidos/<int:id>/status', methods=['PUT'])
@jwt_required()
def avancar_status_pedido(id):
    try:
        data = request.get_json(silent=True) or {}
        novo_status = data.get('status')
        if novo_status not in ('em_compra', 'comprado', 'recebido'):
            return response(False, message='Status inválido. Use: em_compra, comprado, recebido', status_code=400)
        with get_connection() as conn:
            pedido = conn.execute('SELECT id, status FROM pedidos WHERE id = ?', (id,)).fetchone()
            if not pedido:
                return response(False, message='Pedido não encontrado.', status_code=404)

            # Se for recebido, dar entrada no estoque
            if novo_status == 'recebido':
                almoxarifado_id = data.get('almoxarifado_id')
                if not almoxarifado_id:
                    return response(False, message='Informe o almoxarifado de destino.', status_code=400)

                # Buscar itens do pedido
                itens = conn.execute(
                    'SELECT pi.produto_id, pi.quantidade_solicitada FROM pedidos_itens pi WHERE pi.pedido_id = ?',
                    (id,)
                ).fetchall()

                if not itens:
                    return response(False, message='Pedido sem itens para dar entrada.', status_code=400)

                for item in itens:
                    produto_id = item['produto_id']
                    qtd = item['quantidade_solicitada']

                    # Verificar se já existe registro de estoque para este produto+almoxarifado
                    existente = conn.execute(
                        'SELECT id, quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                        (produto_id, almoxarifado_id)
                    ).fetchone()

                    if existente:
                        conn.execute(
                            'UPDATE estoque SET quantidade = quantidade + ?, updated_at = ? WHERE id = ?',
                            (qtd, now_iso(), existente['id'])
                        )
                    else:
                        conn.execute(
                            'INSERT INTO estoque (produto_id, almoxarifado_id, quantidade, estoque_minimo, updated_at) VALUES (?, ?, ?, 0, ?)',
                            (produto_id, almoxarifado_id, qtd, now_iso())
                        )

                    # Registrar movimentação de entrada
                    conn.execute('''
                        INSERT INTO movimentacoes
                            (produto_id, almoxarifado_id, tipo, quantidade,
                             valor_unitario, documento, observacao)
                        VALUES (?, ?, 'entrada', ?, 0, ?, ?)
                    ''', (produto_id, almoxarifado_id, qtd,
                          f'Recebimento Pedido #{id}',
                          f'Entrada por recebimento de Pedido #{id}'))

                current_user = get_jwt_identity()
                user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
                registrar_log(
                    user['id'], current_user, 'receber', 'pedidos',
                    f'Pedido #{id} recebido no almoxarifado #{almoxarifado_id} com {len(itens)} itens',
                    conn=conn
                )

            # Mapa de timestamps
            ts_map = {
                'em_compra': 'data_em_compra',
                'comprado': 'data_comprado',
                'recebido': 'data_recebido'
            }
            coluna_ts = ts_map[novo_status]
            conn.execute(f'UPDATE pedidos SET status = ?, {coluna_ts} = ?, updated_at = ? WHERE id = ?',
                         (novo_status, now_iso(), now_iso(), id))

            if novo_status != 'recebido':
                current_user = get_jwt_identity()
                user = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
                registrar_log(
                    user['id'], current_user, 'atualizar', 'pedidos',
                    f'Pedido #{id} alterado para {novo_status}',
                    conn=conn
                )

            conn.commit()
        return response(True, message=f'Pedido alterado para "{novo_status}".')
    except Exception as e:
        return response(False, message=str(e), status_code=500)
# ============================================================
# TRANSFERENCIAS
# ============================================================

@app.route('/transferencias', methods=['POST'])
@jwt_required()
def criar_transferencia():
    try:
        data = request.get_json(silent=True) or {}
        produto_id = data.get('produto_id')
        almoxarifado_origem_id = data.get('almoxarifado_origem_id')
        almoxarifado_destino_id = data.get('almoxarifado_destino_id')
        quantidade = data.get('quantidade')
        valor_unitario = data.get('valor_unitario')
        documento = data.get('documento')
        tecnico = data.get('tecnico')
        ordem_servico = data.get('ordem_servico')
        observacao = data.get('observacao')

        required = ['produto_id', 'almoxarifado_origem_id', 'almoxarifado_destino_id', 'quantidade']
        for field in required:
            if data.get(field) in (None, ''):
                return response(False, message=f'Campo obrigatório ausente: {field}', status_code=400)

        try:
            quantidade = int(quantidade)
        except (TypeError, ValueError):
            return response(False, message='Quantidade inválida.', status_code=400)

        if almoxarifado_origem_id == almoxarifado_destino_id:
            return response(False, message='Origem e destino devem ser diferentes.', status_code=400)

        if quantidade <= 0:
            return response(False, message='Quantidade deve ser maior que zero.', status_code=400)

        current_user = get_jwt_identity()

        with get_connection() as conn:
            prod = conn.execute('SELECT id, nome FROM produtos WHERE id = ?', (produto_id,)).fetchone()
            if not prod:
                return response(False, message='Produto não encontrado.', status_code=404)

            origem = conn.execute('SELECT id, nome FROM almoxarifados WHERE id = ?', (almoxarifado_origem_id,)).fetchone()
            if not origem:
                return response(False, message='Almoxarifado de origem não encontrado.', status_code=404)

            destino = conn.execute('SELECT id, nome FROM almoxarifados WHERE id = ?', (almoxarifado_destino_id,)).fetchone()
            if not destino:
                return response(False, message='Almoxarifado de destino não encontrado.', status_code=404)

            saldo_row = conn.execute(
                'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                (produto_id, almoxarifado_origem_id)
            ).fetchone()
            saldo_atual = saldo_row['quantidade'] if saldo_row else 0

            if saldo_atual < quantidade:
                return response(False, message='Saldo insuficiente no almoxarifado de origem.', status_code=409)

            conn.execute(
                'UPDATE estoque SET quantidade = quantidade - ? WHERE produto_id = ? AND almoxarifado_id = ?',
                (quantidade, produto_id, almoxarifado_origem_id)
            )

            novo_saldo_origem = saldo_atual - quantidade

            conn.execute(
                '''INSERT INTO transferencias
                   (produto_id, almoxarifado_origem_id, almoxarifado_destino_id,
                    quantidade_total, quantidade_recebida, valor_unitario, status,
                    documento, tecnico, ordem_servico, observacao, data_envio, enviado_por)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (produto_id, almoxarifado_origem_id, almoxarifado_destino_id,
                 quantidade, 0, valor_unitario, 'enviada',
                 documento, tecnico, ordem_servico, observacao, now_iso(), current_user)
            )

            trow = conn.execute('SELECT id FROM transferencias ORDER BY id DESC LIMIT 1').fetchone()
            transfer_id = trow['id']

            conn.execute(
                '''INSERT INTO movimentacoes
                   (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                    valor_unitario, documento, tecnico, ordem_servico, observacao, data)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (produto_id, almoxarifado_origem_id, None, 'saida', quantidade,
                 valor_unitario, documento, tecnico, ordem_servico,
                 f'Transferência #{transfer_id} para {destino["nome"]}', now_iso())
            )

            user_row = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            user_id = user_row['id'] if user_row else None

            registrar_log(user_id, current_user, 'criar', 'transferencias',
                f'Transferência #{transfer_id} enviada: {prod["nome"]} ({quantidade}) de {origem["nome"]} para {destino["nome"]}',
                conn=conn)

            conn.commit()
            return response(True, data={'transfer_id': transfer_id, 'saldo_origem': novo_saldo_origem},
                          message='Transferência enviada com sucesso.')
    except Exception:
        import traceback; traceback.print_exc()
        return response(False, message='Erro interno ao enviar transferência.', status_code=500)

@app.route('/transferencias', methods=['GET'])
@jwt_required()
def listar_transferencias():
    try:
        status_param = request.args.get('status')
        with get_connection() as conn:
            sql = '''SELECT t.*, p.nome AS produto_nome, p.codigo_interno,
                            ao.nome AS almoxarifado_origem_nome,
                            ad.nome AS almoxarifado_destino_nome
                     FROM transferencias t
                     LEFT JOIN produtos p ON t.produto_id = p.id
                     LEFT JOIN almoxarifados ao ON t.almoxarifado_origem_id = ao.id
                     LEFT JOIN almoxarifados ad ON t.almoxarifado_destino_id = ad.id'''
            params = []
            if status_param:
                sql += ' WHERE t.status = ?'
                params.append(status_param)
            sql += ' ORDER BY t.data_envio DESC'
            rows = conn.execute(sql, params).fetchall()
            return response(True, data=rows_to_dict(rows))
    except Exception:
        import traceback; traceback.print_exc()
        return response(False, message='Erro ao listar transferências.', status_code=500)

@app.route('/transferencias/<int:transfer_id>/receber', methods=['POST'])
@jwt_required()
def receber_transferencia(transfer_id):
    try:
        data = request.get_json(silent=True) or {}
        quantidade_receber = data.get('quantidade_receber')

        if quantidade_receber in (None, ''):
            return response(False, message='Campo obrigatório: quantidade_receber', status_code=400)

        try:
            quantidade_receber = int(quantidade_receber)
        except (TypeError, ValueError):
            return response(False, message='Quantidade inválida.', status_code=400)

        if quantidade_receber <= 0:
            return response(False, message='Quantidade deve ser maior que zero.', status_code=400)

        current_user = get_jwt_identity()

        with get_connection() as conn:
            transfer = conn.execute('SELECT * FROM transferencias WHERE id = ?', (transfer_id,)).fetchone()
            if not transfer:
                return response(False, message='Transferência não encontrada.', status_code=404)

            if transfer['status'] not in ('enviada', 'parcial'):
                return response(False, message='Status não permite recebimento.', status_code=400)

            quantidade_total = transfer['quantidade_total']
            quantidade_recebida_atual = transfer['quantidade_recebida'] or 0
            nova_qtd_recebida = quantidade_recebida_atual + quantidade_receber

            if nova_qtd_recebida > quantidade_total:
                return response(False, message='Quantidade excede o total da transferência.', status_code=400)

            produto_id = transfer['produto_id']
            destino_id = transfer['almoxarifado_destino_id']
            origem_id = transfer['almoxarifado_origem_id']
            valor_unitario = transfer['valor_unitario'] if transfer['valor_unitario'] else None

            conn.execute(
                '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                   VALUES (?, ?, ?)
                   ON CONFLICT(produto_id, almoxarifado_id)
                   DO UPDATE SET quantidade = estoque.quantidade + excluded.quantidade''',
                (produto_id, destino_id, quantidade_receber)
            )

            new_status = 'recebida' if nova_qtd_recebida == quantidade_total else 'parcial'

            conn.execute(
                '''UPDATE transferencias SET quantidade_recebida = ?, status = ?, data_recebimento = ? WHERE id = ?''',
                (nova_qtd_recebida, new_status, now_iso(), transfer_id)
            )

            origem = conn.execute('SELECT nome FROM almoxarifados WHERE id = ?', (origem_id,)).fetchone()
            origem_nome = origem['nome'] if origem else ''

            conn.execute(
                '''INSERT INTO movimentacoes
                   (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                    valor_unitario, documento, tecnico, ordem_servico, observacao, data)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (produto_id, destino_id, None, 'entrada', quantidade_receber,
                 valor_unitario, transfer['documento'], transfer['tecnico'],
                 transfer['ordem_servico'],
                 f'Recebimento Transferência #{transfer_id} de {origem_nome}', now_iso())
            )

            user_row = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            user_id = user_row['id'] if user_row else None

            registrar_log(user_id, current_user, 'atualizar', 'transferencias',
                f'Recebimento Transferência #{transfer_id}: +{quantidade_receber} - status {new_status}', conn=conn)

            conn.commit()

            saldo_row = conn.execute(
                'SELECT quantidade FROM estoque WHERE produto_id = ? AND almoxarifado_id = ?',
                (produto_id, destino_id)
            ).fetchone()
            saldo = saldo_row['quantidade'] if saldo_row else 0

            return response(True, data={'saldo_destino': saldo, 'status': new_status,
                                        'quantidade_recebida': nova_qtd_recebida},
                          message='Recebimento registrado com sucesso.')
    except Exception:
        import traceback; traceback.print_exc()
        return response(False, message='Erro ao registrar recebimento.', status_code=500)

@app.route('/transferencias/<int:transfer_id>/rejeitar', methods=['POST'])
@jwt_required()
def rejeitar_transferencia(transfer_id):
    try:
        data = request.get_json(silent=True) or {}
        motivo = data.get('motivo')
        current_user = get_jwt_identity()

        with get_connection() as conn:
            transfer = conn.execute('SELECT * FROM transferencias WHERE id = ?', (transfer_id,)).fetchone()
            if not transfer:
                return response(False, message='Transferência não encontrada.', status_code=404)

            if transfer['status'] not in ('enviada', 'parcial'):
                return response(False, message='Status não permite rejeição.', status_code=400)

            quantidade_total = transfer['quantidade_total']
            quantidade_recebida = transfer['quantidade_recebida'] or 0
            remaining = quantidade_total - quantidade_recebida

            produto_id = transfer['produto_id']
            origem_id = transfer['almoxarifado_origem_id']
            valor_unitario = transfer['valor_unitario'] if transfer['valor_unitario'] else None

            if remaining > 0:
                conn.execute(
                    '''INSERT INTO estoque (produto_id, almoxarifado_id, quantidade)
                       VALUES (?, ?, ?)
                       ON CONFLICT(produto_id, almoxarifado_id)
                       DO UPDATE SET quantidade = estoque.quantidade + excluded.quantidade''',
                    (produto_id, origem_id, remaining)
                )

            obs_atual = transfer['observacao'] or ''
            nova_obs = f'{obs_atual} | Rejeitada: {motivo}' if motivo else (obs_atual or 'Rejeitada')

            conn.execute(
                '''UPDATE transferencias SET status = ?, data_recebimento = ?, observacao = ? WHERE id = ?''',
                ('rejeitada', now_iso(), nova_obs, transfer_id)
            )

            prod = conn.execute('SELECT nome FROM produtos WHERE id = ?', (produto_id,)).fetchone()
            produto_nome = prod['nome'] if prod else ''

            if remaining > 0:
                conn.execute(
                    '''INSERT INTO movimentacoes
                       (produto_id, almoxarifado_id, colaborador_id, tipo, quantidade,
                        valor_unitario, documento, tecnico, ordem_servico, observacao, data)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (produto_id, origem_id, None, 'entrada', remaining,
                     valor_unitario, transfer['documento'], transfer['tecnico'],
                     transfer['ordem_servico'],
                     f'Devolução Transferência #{transfer_id} rejeitada', now_iso())
                )

            user_row = conn.execute('SELECT id FROM usuarios WHERE username = ?', (current_user,)).fetchone()
            user_id = user_row['id'] if user_row else None

            registrar_log(user_id, current_user, 'atualizar', 'transferencias',
                f'Transferência #{transfer_id} rejeitada: {produto_nome} - devolvido {remaining}', conn=conn)

            conn.commit()
            return response(True, message='Transferência rejeitada. Saldo devolvido à origem.')
    except Exception:
        import traceback; traceback.print_exc()
        return response(False, message='Erro ao rejeitar transferência.', status_code=500)


# ============================================================
# EXECUÇÃO
# ============================================================

print("Rotas registradas:")
for rule in app.url_map.iter_rules():
    print(f"  {rule.rule} -> {rule.endpoint}")

init_db()

if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == 'criar-usuario' and len(sys.argv) == 4:
        init_db()
        criar_usuario(sys.argv[2], sys.argv[3])
    else:
        init_db()
        app.run(debug=True, host='127.0.0.1', port=5000)



